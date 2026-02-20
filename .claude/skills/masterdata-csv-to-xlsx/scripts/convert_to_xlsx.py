#!/usr/bin/env python3
"""
マスタデータCSV → XLSX変換スクリプト

指定ディレクトリにあるマスタデータCSVファイルを、sheet_schemaの列定義に準拠した
1つのXLSXファイルに統合します。

使用方法:
    python convert_to_xlsx.py \
        --input-dir <変換元CSVディレクトリ> \
        [--schema-dir projects/glow-masterdata/sheet_schema] \
        [--output-filename 20260220_masterdata.xlsx]
"""

import sys
import argparse
import csv
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    print(json.dumps({
        "error": "openpyxlが見つかりません。`pip install openpyxl` でインストールしてください。"
    }, ensure_ascii=False))
    sys.exit(1)


def parse_sheet_schema(schema_path: Path) -> list[str]:
    """
    sheet_schemaの3行目（ENABLE行）から列定義を取得する。

    Args:
        schema_path: sheet_schemaファイルのパス

    Returns:
        columns: 列名のリスト（"ENABLE"と空列は除外）
    """
    with open(schema_path, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))

    if len(rows) < 3:
        return []

    # 行3（idx=2）: ENABLE行
    enable_row = rows[2]

    # 先頭の "ENABLE" を除外し、空列もスキップ
    columns = [c.strip() for c in enable_row[1:] if c.strip()]
    return columns


def convert_csv_to_sheet(ws, csv_path: Path, schema_columns: list[str]) -> None:
    """
    CSVデータをsheet_schemaの列順でワークシートに書き込む。

    Args:
        ws: openpyxlのワークシートオブジェクト
        csv_path: 変換元CSVファイルのパス
        schema_columns: sheet_schemaから取得した列名リスト
    """
    # 1. ヘッダー行を書き込む（太字）
    ws.append(schema_columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 2. CSVデータを読み込む
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)

    if len(rows) < 3:
        # データなし（ヘッダー3行すらない）
        return

    # 3行目（idx=2）がカラム名行
    csv_header = rows[2]

    # 先頭の "ENABLE" を除いた列名リスト
    if csv_header and csv_header[0].startswith('ENABLE'):
        csv_columns = csv_header[1:]
    else:
        csv_columns = csv_header

    # 4行目以降がデータ行
    data_rows = rows[3:]

    # 3. データをschema_columnsの順番で書き込む
    for data_row in data_rows:
        # CSVの列名とデータを対応付ける辞書を作成
        row_dict = {}
        for i, col_name in enumerate(csv_columns):
            if i < len(data_row):
                row_dict[col_name] = data_row[i]
            else:
                row_dict[col_name] = ''

        # schema_columnsの順番でデータを取得
        # i18n列（例: description.ja）はrow_dictに存在しないので空文字になる
        output_row = [row_dict.get(col, '') for col in schema_columns]
        ws.append(output_row)


def main():
    parser = argparse.ArgumentParser(
        description='マスタデータCSVをsheet_schemaの列定義に従ってXLSXに変換'
    )
    parser.add_argument(
        '--input-dir',
        required=True,
        help='変換元CSVが置かれたディレクトリ'
    )
    parser.add_argument(
        '--schema-dir',
        default='projects/glow-masterdata/sheet_schema',
        help='sheet_schemaのディレクトリ（デフォルト: projects/glow-masterdata/sheet_schema）'
    )
    parser.add_argument(
        '--output-filename',
        default=None,
        help='出力ファイル名（デフォルト: {yyyyMMdd}_masterdata.xlsx）'
    )

    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    schema_dir = Path(args.schema_dir)

    # 入力ディレクトリの存在確認
    if not input_dir.exists():
        print(json.dumps({
            "error": f"入力ディレクトリが見つかりません: {input_dir}"
        }, ensure_ascii=False))
        sys.exit(1)

    # sheet_schemaディレクトリの存在確認
    if not schema_dir.exists():
        print(json.dumps({
            "error": f"sheet_schemaディレクトリが見つかりません: {schema_dir}"
        }, ensure_ascii=False))
        sys.exit(1)

    # 入力CSVファイルを取得（ソート済み）
    csv_files = sorted(input_dir.glob('*.csv'))

    if not csv_files:
        print(json.dumps({
            "warning": f"CSVファイルが見つかりません: {input_dir}",
            "converted": [],
            "skipped": [],
            "summary": {"total": 0, "converted": 0, "skipped": 0}
        }, ensure_ascii=False, indent=2))
        sys.exit(0)

    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    converted = []
    skipped = []

    for csv_file in csv_files:
        schema_file = schema_dir / csv_file.name

        # sheet_schemaが存在しない場合はスキップ
        if not schema_file.exists():
            skipped.append(csv_file.name)
            continue

        # sheet_schemaから列定義を取得
        schema_columns = parse_sheet_schema(schema_file)

        if not schema_columns:
            skipped.append(csv_file.name)
            continue

        # Excelシート名は31文字制限
        sheet_name = csv_file.stem[:31]
        ws = wb.create_sheet(title=sheet_name)

        # CSVデータをシートに書き込む
        try:
            convert_csv_to_sheet(ws, csv_file, schema_columns)
            converted.append(csv_file.name)
        except Exception as e:
            skipped.append(csv_file.name)
            # エラーが発生したシートを削除
            wb.remove(ws)

    # 出力先ディレクトリを作成
    output_dir = input_dir / 'xlsx'
    output_dir.mkdir(parents=True, exist_ok=True)

    # 出力ファイル名の決定
    if args.output_filename:
        filename = args.output_filename
    else:
        filename = f"{datetime.now().strftime('%Y%m%d')}_masterdata.xlsx"

    output_path = output_dir / filename

    # XLSXを保存
    if converted:
        wb.save(output_path)
    else:
        output_path = None

    # サマリーをJSON出力
    result = {
        "output_file": str(output_path) if output_path else None,
        "converted": converted,
        "skipped": skipped,
        "summary": {
            "total": len(csv_files),
            "converted": len(converted),
            "skipped": len(skipped)
        }
    }

    print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0)


if __name__ == "__main__":
    main()
