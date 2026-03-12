#!/usr/bin/env python3
"""
VD全ブロック統合xlsx生成スクリプト

{task-dir}/vd-ingame-design-creator/ 配下の全ブロックの design.json と
generated/*.csv を読み込み、1つの xlsx に統合する。

使用方法:
    python create_all_xlsx.py --task-dir <タスクディレクトリパス>

オプション:
    --template  テンプレートxlsxのパス（デフォルト値あり）
"""

import argparse
import csv
import json
import warnings
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

# CSVテーブルの出力順序（MstEnemyStageParameter は vd_all/data から別途追加）
CSV_TABLE_ORDER = [
    "MstInGame",
    "MstEnemyOutpost",
    "MstKomaLine",
    "MstAutoPlayerSequence",
    "MstPage",
]

# データ行の標準スタイル
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FONT = Font(name="Arial", size=8, color="FF000000")
_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write(ws, coord: str, value, style: bool = True) -> None:
    """セルに値を書き込み、スタイルを適用する。"""
    cell = ws[coord]
    cell.value = value
    if style:
        cell.font = _FONT
        cell.alignment = _ALIGN
        cell.border = _BORDER


def _clear_row(ws, row: int, cols: list) -> None:
    """指定行・列のセル値を空文字でクリアする。"""
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.value = ""


def write_design_to_sheet(ws, data: dict) -> None:
    """design.json のデータをワークシートに書き込む。"""
    block_name = data["block_name"]

    # シート名変更
    ws.title = f"ブロック基礎設計_{block_name}"

    # B1: タイトル
    _write(ws, "B1", data.get("title", f"ブロック基礎設計_{block_name}"), style=False)

    # B4: 要件テキスト
    _write(ws, "B4", data.get("requirements_text", ""))

    # B9〜H9: 基本情報
    bi = data["basic_info"]
    _write(ws, "B9", bi["id"])
    _write(ws, "C9", bi["block_type"])
    _write(ws, "D9", bi["gate_hp"])
    _write(ws, "E9", bi["koma_rows"])
    _write(ws, "F9", bi["group_switch"])
    _write(ws, "G9", bi["sequence_rows"])
    _write(ws, "H9", bi["release_key"])

    # B14〜F16: コマ行（3行固定）
    KOMA_START = 14
    KOMA_COLS = ["B", "C", "D", "E", "F"]
    for i, koma in enumerate(data["koma_rows"]):
        r = KOMA_START + i
        _write(ws, f"B{r}", koma["row"])
        _write(ws, f"C{r}", koma["count"])
        _write(ws, f"D{r}", koma.get("asset", ""))
        _write(ws, f"E{r}", koma["effect"])
        _write(ws, f"F{r}", koma["widths"])
    # テンプレートの余剰行をクリア
    for extra in range(len(data["koma_rows"]), 3):
        _clear_row(ws, KOMA_START + extra, KOMA_COLS)

    # B26〜I??: 登場敵パラメータ
    ENEMY_START = 26
    ENEMY_COLS = ["B", "C", "D", "E", "F", "G", "H", "I"]
    for i, enemy in enumerate(data["enemies"]):
        r = ENEMY_START + i
        _write(ws, f"B{r}", enemy["id"])
        _write(ws, f"C{r}", enemy["name"])
        _write(ws, f"D{r}", enemy["color"])
        _write(ws, f"E{r}", enemy["role"])
        _write(ws, f"F{r}", enemy["hp"])
        _write(ws, f"G{r}", enemy["atk"])
        _write(ws, f"H{r}", enemy["spd"])
        _write(ws, f"I{r}", enemy["desc"])
    # テンプレートのサンプル行をクリア
    TEMPLATE_ENEMY_ROWS = 2
    for extra in range(len(data["enemies"]), TEMPLATE_ENEMY_ROWS):
        _clear_row(ws, ENEMY_START + extra, ENEMY_COLS)

    # B40〜H??: シーケンス構成
    SEQ_START = 40
    SEQ_COLS = ["B", "C", "D", "E", "F", "G", "H"]
    for i, seq in enumerate(data["sequences"]):
        r = SEQ_START + i
        _write(ws, f"B{r}", seq["num"])
        _write(ws, f"C{r}", seq["trigger"])
        _write(ws, f"D{r}", seq["value"])
        _write(ws, f"E{r}", seq["enemy_id"])
        _write(ws, f"F{r}", seq["name"])
        _write(ws, f"G{r}", seq["count"])
        _write(ws, f"H{r}", seq.get("notes", ""))
    # テンプレートのサンプル行をクリア
    TEMPLATE_SEQ_ROWS = 5
    for extra in range(len(data["sequences"]), TEMPLATE_SEQ_ROWS):
        _clear_row(ws, SEQ_START + extra, SEQ_COLS)

    # C58〜C59: ステージ説明文（任意）
    if "stage_description" in data:
        sd = data["stage_description"]
        if sd.get("battle_hint") is not None:
            _write(ws, "C58", sd["battle_hint"])
        if sd.get("stage_text") is not None:
            _write(ws, "C59", sd["stage_text"])
    else:
        ws["C58"].value = ""
        ws["C59"].value = ""


def auto_adjust_column_width(ws, min_width: int = 8, max_width: int = 50) -> None:
    """シートの全列を内容に合わせて幅を自動調整する。"""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def add_csv_rows_to_sheet(wb, sheet_name: str, rows: list, is_first: bool) -> None:
    """CSVデータをワークシートとして追加または既存シートに追記する。

    is_first=True のとき: 新規シートを作成しヘッダー含む全行を書き込む
    is_first=False のとき: 既存シートの末尾にデータ行（ヘッダー除く）を追記する
    """
    if is_first:
        ws = wb.create_sheet(title=sheet_name)
        rows_to_write = rows
        start_row = 1
    else:
        ws = wb[sheet_name]
        rows_to_write = rows[1:]  # ヘッダー行をスキップ
        start_row = ws.max_row + 2  # 1行空けて追記

    for r_offset, row in enumerate(rows_to_write):
        r_idx = start_row + r_offset
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if is_first and r_idx == 1:
                cell.font = Font(name="Arial", size=8, bold=True)
            else:
                cell.font = _FONT


def main() -> None:
    parser = argparse.ArgumentParser(description="VD全ブロック統合xlsx生成")
    parser.add_argument("--task-dir", required=True, help="タスクディレクトリのパス")
    parser.add_argument(
        "--template",
        default=None,
        help="テンプレートxlsxのパス（省略時は task-dir/specs/...）",
    )
    args = parser.parse_args()

    task_dir = Path(args.task_dir)
    design_creator_dir = task_dir / "vd-ingame-design-creator"
    output_dir = task_dir / "vd_all"
    output_path = output_dir / "vd_all.xlsx"

    # テンプレートパスを解決
    if args.template:
        template_path = Path(args.template)
    else:
        template_path = (
            task_dir
            / "specs"
            / "限界チャレンジ(VD)_アウトゲーム関連_ブロック基礎設計テンプレート.xlsx"
        )

    if not template_path.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {template_path}")

    if not design_creator_dir.exists():
        raise FileNotFoundError(
            f"vd-ingame-design-creator ディレクトリが見つかりません: {design_creator_dir}"
        )

    # ブロックフォルダを探索（vd_all 除く、名前でソート）
    block_dirs = sorted(
        [d for d in design_creator_dir.iterdir() if d.is_dir() and d.name != "vd_all"]
    )

    if not block_dirs:
        raise ValueError(f"ブロックフォルダが見つかりません: {design_creator_dir}")

    print(f"検出されたブロック: {[d.name for d in block_dirs]}")

    # 出力ディレクトリを作成
    output_dir.mkdir(parents=True, exist_ok=True)

    # テンプレートをロード（テンプレートシートは後で削除）
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(template_path)
    template_ws = wb.active

    # ── 1. 各ブロックのブロック基礎設計シートを作成 ──────────────────
    print("\n[ブロック基礎設計シート作成]")
    for block_dir in block_dirs:
        design_json_path = block_dir / "design.json"
        if not design_json_path.exists():
            print(f"  警告: design.json が見つかりません: {design_json_path}")
            continue

        with open(design_json_path, encoding="utf-8") as f:
            data = json.load(f)

        block_name = data["block_name"]
        print(f"  作成: ブロック基礎設計_{block_name}")

        # テンプレートシートをコピーして新しいシートを作成
        ws = wb.copy_worksheet(template_ws)
        write_design_to_sheet(ws, data)

    # テンプレートシートを削除（全ブロックコピー後）
    wb.remove(template_ws)

    # ── 2. 全ブロックのCSVシートをテーブル名ごとにまとめて追加 ─────────
    print("\n[CSVシート作成]")
    for table_name in CSV_TABLE_ORDER:
        is_first_block = True
        for block_dir in block_dirs:
            csv_path = block_dir / "generated" / f"{table_name}.csv"
            if not csv_path.exists():
                continue

            with open(csv_path, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)

            if not rows:
                continue

            action = "新規作成" if is_first_block else "追記"
            print(f"  {action}: {table_name} ← {block_dir.name}")
            add_csv_rows_to_sheet(wb, table_name, rows, is_first=is_first_block)
            is_first_block = False

    # ── 3. MstEnemyStageParameter を vd_all/data から追加 ──────────────
    print("\n[MstEnemyStageParameter シート作成]")
    esp_path = design_creator_dir / "vd_all" / "data" / "MstEnemyStageParameter.csv"
    if esp_path.exists():
        with open(esp_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if rows:
            print(f"  作成: MstEnemyStageParameter ← vd_all/data")
            ws_esp = wb.create_sheet(title="MstEnemyStageParameter")
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws_esp.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(name="Arial", size=8, bold=True)
                    else:
                        cell.font = _FONT
    else:
        print(f"  警告: MstEnemyStageParameter.csv が見つかりません: {esp_path}")

    # ── 4. 全テーブルシートの列幅を自動調整 ────────────────────────────
    print("\n[列幅自動調整]")
    for sheet_name in CSV_TABLE_ORDER:
        if sheet_name in wb.sheetnames:
            auto_adjust_column_width(wb[sheet_name])
            print(f"  調整: {sheet_name}")

    esp_sheet_name = "MstEnemyStageParameter"
    if esp_sheet_name in wb.sheetnames:
        auto_adjust_column_width(wb[esp_sheet_name])
        print(f"  調整: {esp_sheet_name}")

    # ── 5. 保存 ────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\n生成完了: {output_path}")
    print(f"シート一覧: {wb.sheetnames}")


if __name__ == "__main__":
    main()
