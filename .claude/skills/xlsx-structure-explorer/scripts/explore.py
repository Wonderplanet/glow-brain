#!/usr/bin/env python3
"""
XLSX構造調査スクリプト

使用例:
  python3 explore.py --mode overview file.xlsx
  python3 explore.py --sheet 1話 --mode cells --row-range 28-35 --col-range BE-CV file.xlsx
  python3 explore.py --sheet 通常ブロック --mode validations file.xlsx
  python3 explore.py --mode compare --sheet 1話 --files a.xlsx,b.xlsx,c.xlsx
"""

import sys
import argparse
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def parse_col_range(col_range: str) -> tuple[int, int]:
    """列範囲をパース。例: 'BE-CV' → (57, 100), '57-100' → (57, 100)"""
    parts = col_range.split("-")
    if len(parts) != 2:
        raise ValueError(f"列範囲の形式が不正です: {col_range} (例: BE-CV または 57-100)")
    start, end = parts
    try:
        col_start = int(start)
    except ValueError:
        col_start = column_index_from_string(start)
    try:
        col_end = int(end)
    except ValueError:
        col_end = column_index_from_string(end)
    return col_start, col_end


def parse_row_range(row_range: str) -> tuple[int, int]:
    """行範囲をパース。例: '28-35' → (28, 35)"""
    parts = row_range.split("-")
    if len(parts) != 2:
        raise ValueError(f"行範囲の形式が不正です: {row_range} (例: 28-35)")
    return int(parts[0]), int(parts[1])


def format_cell_value(cell, data_only: bool = False) -> str:
    """セル値を表示用文字列に変換"""
    if cell.value is None:
        return ""
    val = str(cell.value)
    # 数式かどうかを判定
    if isinstance(cell.value, str) and cell.value.startswith("="):
        if len(val) > 50:
            return val[:47] + "..."
        return val
    if len(val) > 40:
        return val[:37] + "..."
    return val


# ============================================================
# モード: overview
# ============================================================

def mode_overview(wb: openpyxl.Workbook, xlsx_path: str) -> str:
    lines = [f"# XLSX構造レポート: {xlsx_path}", ""]
    lines.append("## シート一覧")
    lines.append("| # | シート名 | 最終行 | 最終列 |")
    lines.append("|---|---------|--------|--------|")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        col_label = f"{get_column_letter(max_col)}({max_col})" if max_col > 0 else "0"
        lines.append(f"| {i} | {name} | {max_row} | {col_label} |")
    return "\n".join(lines)


# ============================================================
# モード: cells
# ============================================================

def mode_cells(ws, sheet_name: str, row_range: tuple[int, int] | None,
               col_range: tuple[int, int] | None, data_only: bool) -> str:
    r_start = row_range[0] if row_range else 1
    r_end = row_range[1] if row_range else min(ws.max_row or 1, 50)
    c_start = col_range[0] if col_range else 1
    c_end = col_range[1] if col_range else min(ws.max_column or 1, 26)

    col_labels = [get_column_letter(c) for c in range(c_start, c_end + 1)]
    range_desc = f"{get_column_letter(c_start)}{r_start}〜{get_column_letter(c_end)}{r_end}"

    lines = [f"## セル範囲: {sheet_name}シート {range_desc}", ""]

    # ヘッダー行
    header = "| 行 | " + " | ".join(col_labels) + " |"
    sep = "|" + "---|" * (len(col_labels) + 1)
    lines.append(header)
    lines.append(sep)

    for row in range(r_start, r_end + 1):
        cells = []
        for col in range(c_start, c_end + 1):
            cell = ws.cell(row=row, column=col)
            cells.append(format_cell_value(cell, data_only))
        lines.append(f"| {row} | " + " | ".join(cells) + " |")

    return "\n".join(lines)


# ============================================================
# モード: formulas
# ============================================================

def mode_formulas(ws, sheet_name: str, row_range: tuple[int, int] | None,
                  col_range: tuple[int, int] | None) -> str:
    r_start = row_range[0] if row_range else 1
    r_end = row_range[1] if row_range else (ws.max_row or 1)
    c_start = col_range[0] if col_range else 1
    c_end = col_range[1] if col_range else (ws.max_column or 1)

    lines = [f"## 数式一覧: {sheet_name}シート", ""]
    lines.append("| セル | 数式 |")
    lines.append("|------|------|")

    found = 0
    for row in range(r_start, r_end + 1):
        for col in range(c_start, c_end + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                addr = f"{get_column_letter(col)}{row}"
                formula = cell.value if len(cell.value) <= 80 else cell.value[:77] + "..."
                lines.append(f"| {addr} | `{formula}` |")
                found += 1

    if found == 0:
        lines.append("| (数式なし) | - |")
    else:
        lines.append("")
        lines.append(f"合計 {found} 個の数式")

    return "\n".join(lines)


# ============================================================
# モード: merges
# ============================================================

def mode_merges(ws, sheet_name: str) -> str:
    lines = [f"## 結合セル一覧: {sheet_name}シート", ""]
    merges = list(ws.merged_cells.ranges)
    if not merges:
        lines.append("結合セルなし")
        return "\n".join(lines)

    lines.append("| # | 範囲 | 開始行 | 開始列 | 行数 | 列数 |")
    lines.append("|---|------|--------|--------|------|------|")
    for i, merge in enumerate(sorted(merges, key=lambda m: (m.min_row, m.min_col)), 1):
        row_span = merge.max_row - merge.min_row + 1
        col_span = merge.max_col - merge.min_col + 1
        range_str = str(merge)
        lines.append(f"| {i} | {range_str} | {merge.min_row} | {get_column_letter(merge.min_col)} | {row_span} | {col_span} |")

    lines.append("")
    lines.append(f"合計 {len(merges)} 個の結合セル")
    return "\n".join(lines)


# ============================================================
# モード: validations
# ============================================================

def mode_validations(ws, sheet_name: str) -> str:
    lines = [f"## データバリデーション: {sheet_name}シート", ""]

    if not ws.data_validations.dataValidation:
        lines.append("バリデーションなし")
        return "\n".join(lines)

    lines.append("| # | 範囲 | 種類 | 数式1（入力値/リスト） | プロンプト |")
    lines.append("|---|------|------|----------------------|------------|")

    for i, dv in enumerate(ws.data_validations.dataValidation, 1):
        sqref = str(dv.sqref)
        dv_type = dv.type or ""
        formula1 = dv.formula1 or ""
        if len(formula1) > 60:
            formula1 = formula1[:57] + "..."
        prompt = dv.promptTitle or ""
        lines.append(f"| {i} | {sqref} | {dv_type} | `{formula1}` | {prompt} |")

    lines.append("")
    lines.append(f"合計 {len(ws.data_validations.dataValidation)} 個のバリデーション")
    return "\n".join(lines)


# ============================================================
# モード: compare
# ============================================================

def mode_compare(files: list[str], sheet_name: str,
                 row_range: tuple[int, int] | None,
                 col_range: tuple[int, int] | None) -> str:
    if not files:
        return "compareモードには --files オプションが必要です"

    lines = [f"## ファイル比較: {sheet_name}シート", ""]
    workbooks = []
    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=False)
            workbooks.append((f, wb))
        except Exception as e:
            lines.append(f"- {f}: 読み込みエラー ({e})")

    if not workbooks:
        return "\n".join(lines)

    # 先頭ファイルの範囲を基準にする
    _, wb_ref = workbooks[0]
    if sheet_name not in wb_ref.sheetnames:
        return f"シート '{sheet_name}' が {files[0]} に存在しません"

    ws_ref = wb_ref[sheet_name]
    r_start = row_range[0] if row_range else 1
    r_end = row_range[1] if row_range else min(ws_ref.max_row or 1, 20)
    c_start = col_range[0] if col_range else 1
    c_end = col_range[1] if col_range else min(ws_ref.max_column or 1, 10)

    col_labels = [get_column_letter(c) for c in range(c_start, c_end + 1)]

    for f, wb in workbooks:
        lines.append(f"### {f}")
        if sheet_name not in wb.sheetnames:
            lines.append(f"シート '{sheet_name}' が存在しません")
            lines.append("")
            continue
        ws = wb[sheet_name]
        header = "| 行 | " + " | ".join(col_labels) + " |"
        sep = "|" + "---|" * (len(col_labels) + 1)
        lines.append(header)
        lines.append(sep)
        for row in range(r_start, r_end + 1):
            cells = []
            for col in range(c_start, c_end + 1):
                cell = ws.cell(row=row, column=col)
                cells.append(format_cell_value(cell))
            lines.append(f"| {row} | " + " | ".join(cells) + " |")
        lines.append("")

    return "\n".join(lines)


# ============================================================
# メイン
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="XLSXファイル構造を調査してMarkdownレポートを出力する",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  # シート一覧確認
  python3 explore.py --mode overview file.xlsx

  # セル範囲確認
  python3 explore.py --sheet 1話 --mode cells --row-range 28-35 --col-range BE-CV file.xlsx

  # 数式のみ抽出
  python3 explore.py --sheet 1話 --mode formulas --row-range 1-50 file.xlsx

  # 結合セル一覧
  python3 explore.py --sheet 1話 --mode merges file.xlsx

  # プルダウン（データバリデーション）一覧
  python3 explore.py --sheet 通常ブロック --mode validations file.xlsx

  # 複数ファイル比較
  python3 explore.py --mode compare --sheet 1話 --row-range 1-10 --col-range A-E --files a.xlsx,b.xlsx,c.xlsx dummy.xlsx
        """
    )
    parser.add_argument("xlsx_path", help="調査対象のXLSXファイルパス（compareモードではダミー可）")
    parser.add_argument("--sheet", "-s", help="シート名（省略時: 全シート対象 or overview）")
    parser.add_argument(
        "--mode", "-m",
        choices=["overview", "cells", "formulas", "merges", "validations", "compare"],
        default="overview",
        help="調査モード (default: overview)"
    )
    parser.add_argument("--row-range", help="行範囲 例: 28-35")
    parser.add_argument("--col-range", help="列範囲 例: BE-CV または 57-100")
    parser.add_argument("--files", help="compareモード: カンマ区切りのXLSXファイルリスト")
    parser.add_argument("--data-only", action="store_true", help="キャッシュ値を使用（数式ではなく計算結果を表示）")

    args = parser.parse_args()

    # 範囲パース
    row_range = parse_row_range(args.row_range) if args.row_range else None
    col_range = parse_col_range(args.col_range) if args.col_range else None

    # compareモード
    if args.mode == "compare":
        if not args.files:
            print("compareモードには --files オプションが必要です", file=sys.stderr)
            sys.exit(1)
        if not args.sheet:
            print("compareモードには --sheet オプションが必要です", file=sys.stderr)
            sys.exit(1)
        file_list = [f.strip() for f in args.files.split(",")]
        print(mode_compare(file_list, args.sheet, row_range, col_range))
        return

    # ファイル読み込み
    try:
        wb = openpyxl.load_workbook(args.xlsx_path, data_only=args.data_only)
    except FileNotFoundError:
        print(f"エラー: ファイルが見つかりません: {args.xlsx_path}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"エラー: ファイル読み込み失敗: {e}", file=sys.stderr)
        sys.exit(1)

    # overviewモードはシート指定不要
    if args.mode == "overview":
        print(mode_overview(wb, args.xlsx_path))
        return

    # シート指定が必要なモード
    if not args.sheet:
        print(f"エラー: --modeが'{args.mode}'の場合は --sheet が必要です", file=sys.stderr)
        print(f"利用可能なシート: {', '.join(wb.sheetnames)}", file=sys.stderr)
        sys.exit(1)

    if args.sheet not in wb.sheetnames:
        print(f"エラー: シート '{args.sheet}' が存在しません", file=sys.stderr)
        print(f"利用可能なシート: {', '.join(wb.sheetnames)}", file=sys.stderr)
        sys.exit(1)

    ws = wb[args.sheet]

    if args.mode == "cells":
        print(mode_cells(ws, args.sheet, row_range, col_range, args.data_only))
    elif args.mode == "formulas":
        print(mode_formulas(ws, args.sheet, row_range, col_range))
    elif args.mode == "merges":
        print(mode_merges(ws, args.sheet))
    elif args.mode == "validations":
        print(mode_validations(ws, args.sheet))


if __name__ == "__main__":
    main()
