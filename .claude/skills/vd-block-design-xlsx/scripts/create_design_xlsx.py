#!/usr/bin/env python3
"""
VDブロック基礎設計xlsx生成スクリプト

テンプレートxlsxをコピーし、Claudeが渡したJSONデータを書き込んで
{block_name}_design.xlsx として出力する。

使用方法:
    python create_design_xlsx.py --block-dir <パス> --data-json '<JSON>'

オプション:
    --template  テンプレートxlsxのパス（デフォルト値あり）
"""

import argparse
import json
import shutil
import warnings
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

# テンプレートxlsxのデフォルトパス（リポジトリルートからの相対パス）
TEMPLATE_PATH = (
    "domain/tasks/20260310_115400_vd_ingame_masterdata_generation/"
    "specs/限界チャレンジ(VD)_アウトゲーム関連_ブロック基礎設計テンプレート.xlsx"
)

# データ行の標準スタイル
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FONT = Font(name="Arial", size=8, color="FF000000")
_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write(ws, coord: str, value, style: bool = True) -> None:
    """セルに値を書き込み、スタイルを適用する。"""
    cell = ws[coord]
    cell.value = value
    if style:
        cell.font = _FONT
        cell.alignment = _ALIGN
        cell.border = _BORDER


def _clear_row(ws, row: int, cols: list[str]) -> None:
    """指定行・列のセル値を空文字でクリアする。"""
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.value = ""


def main() -> None:
    parser = argparse.ArgumentParser(description="VDブロック基礎設計xlsx生成")
    parser.add_argument("--block-dir", required=True, help="ブロックフォルダのパス")
    parser.add_argument("--data-json", required=True, help="入力データJSON文字列")
    parser.add_argument(
        "--template", default=TEMPLATE_PATH, help="テンプレートxlsxのパス"
    )
    args = parser.parse_args()

    data = json.loads(args.data_json)
    block_dir = Path(args.block_dir)
    block_name = data["block_name"]
    output_path = block_dir / f"{block_name}_design.xlsx"

    # テンプレートをコピー
    shutil.copy2(args.template, output_path)

    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # シート名変更
    ws.title = f"ブロック基礎設計_{block_name}"

    # ── B1: タイトル ──────────────────────────────────────
    _write(ws, "B1", data["title"])

    # ── B4: 要件テキスト ──────────────────────────────────
    _write(ws, "B4", data["requirements_text"])

    # ── B9〜H9: 基本情報 ─────────────────────────────────
    bi = data["basic_info"]
    _write(ws, "B9", bi["id"])
    _write(ws, "C9", bi["block_type"])
    _write(ws, "D9", bi["gate_hp"])
    _write(ws, "E9", bi["koma_rows"])
    _write(ws, "F9", bi["group_switch"])
    _write(ws, "G9", bi["sequence_rows"])
    _write(ws, "H9", bi["release_key"])

    # ── B14〜F16: コマ行（3行固定）────────────────────────
    KOMA_START = 14
    KOMA_COLS = ["B", "C", "D", "E", "F"]
    for i, koma in enumerate(data["koma_rows"]):
        r = KOMA_START + i
        _write(ws, f"B{r}", koma["row"])
        _write(ws, f"C{r}", koma["count"])
        _write(ws, f"D{r}", koma.get("asset", ""))
        _write(ws, f"E{r}", koma["effect"])
        _write(ws, f"F{r}", koma["widths"])
    # テンプレートの余剰行をクリア
    for extra in range(len(data["koma_rows"]), 3):
        _clear_row(ws, KOMA_START + extra, KOMA_COLS)

    # ── B26〜I??: 登場敵パラメータ ──────────────────────
    ENEMY_START = 26
    ENEMY_COLS = ["B", "C", "D", "E", "F", "G", "H", "I"]
    for i, enemy in enumerate(data["enemies"]):
        r = ENEMY_START + i
        _write(ws, f"B{r}", enemy["id"])
        _write(ws, f"C{r}", enemy["name"])
        _write(ws, f"D{r}", enemy["color"])
        _write(ws, f"E{r}", enemy["role"])
        _write(ws, f"F{r}", enemy["hp"])
        _write(ws, f"G{r}", enemy["atk"])
        _write(ws, f"H{r}", enemy["spd"])
        _write(ws, f"I{r}", enemy["desc"])
    # テンプレートに既存のサンプル行が残っている場合はクリア
    TEMPLATE_ENEMY_ROWS = 2  # テンプレートのサンプル行数
    for extra in range(len(data["enemies"]), TEMPLATE_ENEMY_ROWS):
        _clear_row(ws, ENEMY_START + extra, ENEMY_COLS)

    # ── B40〜H??: シーケンス構成 ─────────────────────────
    SEQ_START = 40
    SEQ_COLS = ["B", "C", "D", "E", "F", "G", "H"]
    for i, seq in enumerate(data["sequences"]):
        r = SEQ_START + i
        _write(ws, f"B{r}", seq["num"])
        _write(ws, f"C{r}", seq["trigger"])
        _write(ws, f"D{r}", seq["value"])
        _write(ws, f"E{r}", seq["enemy_id"])
        _write(ws, f"F{r}", seq["name"])
        _write(ws, f"G{r}", seq["count"])
        _write(ws, f"H{r}", seq.get("notes", ""))
    # テンプレートに既存のサンプル行が残っている場合はクリア
    TEMPLATE_SEQ_ROWS = 5  # テンプレートのサンプル行数
    for extra in range(len(data["sequences"]), TEMPLATE_SEQ_ROWS):
        _clear_row(ws, SEQ_START + extra, SEQ_COLS)

    # ── C58〜C59: ステージ説明文（任意）───────────────────
    if "stage_description" in data:
        sd = data["stage_description"]
        if sd.get("battle_hint") is not None:
            _write(ws, "C58", sd["battle_hint"])
        if sd.get("stage_text") is not None:
            _write(ws, "C59", sd["stage_text"])
    else:
        # テンプレートのサンプルテキストをクリア
        ws["C58"].value = ""
        ws["C59"].value = ""

    wb.save(output_path)
    print(f"生成完了: {output_path}")


if __name__ == "__main__":
    main()
