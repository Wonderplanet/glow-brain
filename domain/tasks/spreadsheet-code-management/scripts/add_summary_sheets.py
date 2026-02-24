#!/usr/bin/env python3
"""
管理シート（サマリー出力設定）ドリブンのサマリーシート生成スクリプト

「サマリー出力設定」シートの設定に基づいて、各テーブルのサマリーシートを動的に生成する。

シートが存在しない場合は、デフォルト設定（MstPage・MstKomaLine）で自動作成する。

各話シートのデータ位置（既存スクリプト add_koma_masterdata_output.py で設定済み）:
  - MstPage: 行31のみ  / EB(132)=id, EC(133)=release_key
  - MstKomaLine: 行31〜35 / EF(136)=id 〜 FU(177)=release_key

Usage:
    python3 scripts/add_summary_sheets.py
"""

import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


# ---- 定数 -------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent
BASE_DIR = SCRIPT_DIR.parent

# 対象XLSXパス（読み込み・上書き保存）
TARGET_XLSX = BASE_DIR / "xlsx" / "【ストーリー】必ず生きて帰る_コマ設計出力追加.xlsx"

# 管理シート名
CONFIG_SHEET_NAME = "サマリー出力設定"

# スタイル
HEADER_FILL_COLOR = "FFFFF2CC"  # 薄い黄色（各話シートと統一）
HEADER_FILL = PatternFill(patternType="solid", fgColor=HEADER_FILL_COLOR)
LINK_FONT = Font(color="0563C1", underline="single")  # ハイパーリンク文字スタイル

# 管理シートの列ヘッダー
CONFIG_HEADERS = ["テーブル名", "ソースシート", "開始列", "終了列", "開始行", "終了行"]

# 管理シートの初期データ（現在のハードコード値を移行）
DEFAULT_CONFIG = [
    {"table_name": "MstPage",     "source_sheet": "1話", "start_col": "EB", "end_col": "EC", "start_row": 31, "end_row": 31},
    {"table_name": "MstPage",     "source_sheet": "2話", "start_col": "EB", "end_col": "EC", "start_row": 31, "end_row": 31},
    {"table_name": "MstPage",     "source_sheet": "3話", "start_col": "EB", "end_col": "EC", "start_row": 31, "end_row": 31},
    {"table_name": "MstPage",     "source_sheet": "4話", "start_col": "EB", "end_col": "EC", "start_row": 31, "end_row": 31},
    {"table_name": "MstPage",     "source_sheet": "5話", "start_col": "EB", "end_col": "EC", "start_row": 31, "end_row": 31},
    {"table_name": "MstPage",     "source_sheet": "6話", "start_col": "EB", "end_col": "EC", "start_row": 31, "end_row": 31},
    {"table_name": "MstKomaLine", "source_sheet": "1話", "start_col": "EF", "end_col": "FU", "start_row": 31, "end_row": 35},
    {"table_name": "MstKomaLine", "source_sheet": "2話", "start_col": "EF", "end_col": "FU", "start_row": 31, "end_row": 35},
    {"table_name": "MstKomaLine", "source_sheet": "3話", "start_col": "EF", "end_col": "FU", "start_row": 31, "end_row": 35},
    {"table_name": "MstKomaLine", "source_sheet": "4話", "start_col": "EF", "end_col": "FU", "start_row": 31, "end_row": 35},
    {"table_name": "MstKomaLine", "source_sheet": "5話", "start_col": "EF", "end_col": "FU", "start_row": 31, "end_row": 35},
    {"table_name": "MstKomaLine", "source_sheet": "6話", "start_col": "EF", "end_col": "FU", "start_row": 31, "end_row": 35},
]


# ---- 管理シート操作 ----------------------------------------------------------

def ensure_config_sheet(wb: openpyxl.Workbook) -> None:
    """
    「サマリー出力設定」シートが存在しない場合のみ、初期データ付きで新規作成する。
    存在する場合はスキップ（ユーザーの変更を尊重）。
    """
    if CONFIG_SHEET_NAME in wb.sheetnames:
        print(f"  → 「{CONFIG_SHEET_NAME}」シートが既に存在します（スキップ）")
        return

    ws = wb.create_sheet(CONFIG_SHEET_NAME)

    # ヘッダー行（行1）
    for col_idx, header in enumerate(CONFIG_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL

    # 初期データ（12行）
    for row_idx, entry in enumerate(DEFAULT_CONFIG, start=2):
        ws.cell(row=row_idx, column=1).value = entry["table_name"]
        ws.cell(row=row_idx, column=2).value = entry["source_sheet"]
        ws.cell(row=row_idx, column=3).value = entry["start_col"]
        ws.cell(row=row_idx, column=4).value = entry["end_col"]
        ws.cell(row=row_idx, column=5).value = entry["start_row"]
        ws.cell(row=row_idx, column=6).value = entry["end_row"]

    print(f"  ✓ 「{CONFIG_SHEET_NAME}」シートを初期データ（{len(DEFAULT_CONFIG)}行）で新規作成")


def load_config_from_sheet(wb: openpyxl.Workbook) -> list[dict]:
    """
    「サマリー出力設定」シートを読み込み、設定エントリのリストを返す。
    空行・テーブル名なし行はスキップ。
    """
    ws = wb[CONFIG_SHEET_NAME]
    entries = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        table_name, source_sheet, start_col, end_col, start_row, end_row = row[:6]

        # テーブル名が空の行はスキップ
        if not table_name:
            continue

        try:
            entries.append({
                "table_name": str(table_name).strip(),
                "source_sheet": str(source_sheet).strip(),
                "start_col": str(start_col).strip().upper(),
                "end_col": str(end_col).strip().upper(),
                "start_row": int(start_row),
                "end_row": int(end_row),
            })
        except (ValueError, TypeError) as e:
            print(f"  警告: 設定行を読み飛ばしました（{row}）: {e}", file=sys.stderr)

    return entries


# ---- サマリーシート生成 -------------------------------------------------------

def read_header_row(
    wb: openpyxl.Workbook,
    source_sheet: str,
    start_col: str,
    end_col: str,
    header_row: int,
) -> list[str]:
    """
    ソースシートの指定行からヘッダー文字列のリストを読み取る。
    header_row <= 0 の場合（start_row == 1 のとき）は列文字をそのままヘッダーとして返す。
    ソースシートが存在しない場合も列文字をフォールバックとして使う。
    """
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    col_letters = [get_column_letter(i) for i in range(start_idx, end_idx + 1)]

    # start_row == 1 の場合はフォールバック：列文字をヘッダーとして使う
    if header_row <= 0:
        return col_letters

    if source_sheet not in wb.sheetnames:
        return col_letters

    ws = wb[source_sheet]
    headers = []
    for col_idx, col_letter in zip(range(start_idx, end_idx + 1), col_letters):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        headers.append(str(cell_value) if cell_value is not None else col_letter)

    return headers


def add_summary_sheet(
    wb: openpyxl.Workbook,
    table_name: str,
    entries: list[dict],
) -> None:
    """
    テーブルのサマリーシートを生成する。

    レイアウト:
      行1: ヘッダー（A=ソースシート, B以降=ソースシートのヘッダー行から取得）
      エントリごとに start_row ~ end_row の各行を追加
      A列: 各エントリの最初の行のみハイパーリンク
      エントリ間: 空行1行を挿入
    """
    # 既存シートがあれば削除して再作成
    if table_name in wb.sheetnames:
        del wb[table_name]
    ws = wb.create_sheet(table_name)

    # ヘッダーを最初のエントリのソースシートから取得
    first_entry = entries[0]
    header_row_num = first_entry["start_row"] - 1  # ヘッダーはデータ行の1つ上
    headers = read_header_row(
        wb,
        first_entry["source_sheet"],
        first_entry["start_col"],
        first_entry["end_col"],
        header_row_num,
    )

    # ヘッダー行（行1）
    all_headers = ["ソースシート"] + headers
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL

    # 各エントリのデータ行
    current_row = 2

    for entry in entries:
        source_sheet = entry["source_sheet"]
        start_row = entry["start_row"]
        end_row = entry["end_row"]
        entry_start_col_idx = column_index_from_string(entry["start_col"])
        entry_end_col_idx = column_index_from_string(entry["end_col"])

        for data_row in range(start_row, end_row + 1):
            is_first_row = data_row == start_row

            # A列: 最初の行にのみハイパーリンクを設定
            if is_first_row:
                link_cell = ws.cell(row=current_row, column=1)
                link_cell.value = source_sheet
                link_cell.hyperlink = f"#'{source_sheet}'!A1"
                link_cell.font = LINK_FONT

            # B列以降: ソースシートへの参照式
            for data_col_idx, src_col_num in enumerate(
                range(entry_start_col_idx, entry_end_col_idx + 1), start=2
            ):
                src_col_letter = get_column_letter(src_col_num)
                ws.cell(row=current_row, column=data_col_idx).value = (
                    f"='{source_sheet}'!{src_col_letter}{data_row}"
                )

            current_row += 1

        # エントリ間に空行を挿入
        current_row += 1

    row_count_per_entry = first_entry["end_row"] - first_entry["start_row"] + 1
    print(
        f"  ✓ {table_name}シート追加（{len(entries)}エントリ × {row_count_per_entry}行）"
    )


# ---- メイン ------------------------------------------------------------------

def main() -> None:
    target_path = TARGET_XLSX.resolve()

    if not target_path.exists():
        print(f"エラー: XLSXが見つかりません: {target_path}", file=sys.stderr)
        sys.exit(1)

    print(f"対象: {target_path}")

    # [1] XLSXを読み込む（data_only=False で式を保持）
    print("\n[1/4] XLSXを読み込み中...")
    wb = openpyxl.load_workbook(str(target_path), data_only=False)

    # [2] 管理シートを確認/作成
    print(f"\n[2/4] 「{CONFIG_SHEET_NAME}」シートを確認中...")
    ensure_config_sheet(wb)

    # [3] 設定を読み込み、テーブルごとにサマリーシートを生成
    print("\n[3/4] サマリーシート生成中...")
    config_entries = load_config_from_sheet(wb)

    if not config_entries:
        print("エラー: 設定エントリが見つかりません。処理を中断します。", file=sys.stderr)
        sys.exit(1)

    # 出現順を保持しながらテーブル名でグループ化
    table_groups: OrderedDict[str, list[dict]] = OrderedDict()
    for entry in config_entries:
        table_name = entry["table_name"]
        if table_name not in table_groups:
            table_groups[table_name] = []
        table_groups[table_name].append(entry)

    # 各テーブルのサマリーシートを生成
    for table_name, entries in table_groups.items():
        add_summary_sheet(wb, table_name, entries)

    # [4] 上書き保存
    print(f"\n[4/4] XLSX保存中: {target_path}")
    wb.save(str(target_path))

    size_kb = target_path.stat().st_size / 1024
    print(f"\n完了!")
    print(f"  ファイルサイズ: {size_kb:,.1f} KB")
    print(f"  出力先: {target_path}")
    print()
    print("追加されたシート:")
    for table_name, entries in table_groups.items():
        first = entries[0]
        row_count = first["end_row"] - first["start_row"] + 1
        print(f"  ・{table_name:<15} - {len(entries)}エントリ × {row_count}行")
    print()
    print("A列のシート名をクリックすると対応する話シートへジャンプできます。")


if __name__ == "__main__":
    main()
