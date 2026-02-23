#!/usr/bin/env python3
"""
CSV / JSON → XLSX 再構築スクリプト

xlsx_to_csv_json.py で生成した outputs/<name>/ から XLSX を再構築する。
元の XLSX（xlsx/ フォルダ）は一切読まない・書かない。

出力先:
  outputs/<name>/<name>.xlsx   （--output で任意パス指定も可能）

Usage:
    python3 scripts/csv_json_to_xlsx.py outputs/【ストーリー】必ず生きて帰る
    python3 scripts/csv_json_to_xlsx.py outputs/【ストーリー】必ず生きて帰る --output my.xlsx
"""

import sys
import csv
import json
import argparse
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.formatting.rule import (
    Rule,
    ColorScale,
    DataBar,
    IconSet,
    FormatObject,
    DifferentialStyle,
)


# ---- スタイル復元ヘルパー ---------------------------------------------------

def _make_color(c) -> Color | None:
    """
    styles.json のカラー値から openpyxl の Color オブジェクトを作成。

    - 文字列 → Color(rgb=...)
    - {"theme": N, "tint": T} → Color(theme=N, tint=T)
    - {"indexed": N} → Color(indexed=N)
    """
    if c is None:
        return None
    if isinstance(c, str):
        return Color(rgb=c)
    if isinstance(c, dict):
        if 'theme' in c:
            return Color(theme=c['theme'], tint=c.get('tint', 0.0))
        if 'indexed' in c:
            return Color(indexed=c['indexed'])
    return None


def _make_side(d: dict | None) -> Side:
    """styles.json の罫線辺データから openpyxl の Side オブジェクトを作成。"""
    if not d:
        return Side()
    color = _make_color(d.get('color'))
    return Side(
        border_style=d.get('border_style'),
        color=color,
    )


# ---- style_map 構築 ---------------------------------------------------------

def build_style_map(styles_json: Path) -> dict[str, dict[str, dict]]:
    """
    styles.json を読み込み、{ シート名: { 座標: スタイルdict } } を返す。
    ファイルが存在しない場合は {} を返す（後方互換）。
    """
    if not styles_json.exists():
        return {}

    with open(styles_json, encoding='utf-8') as f:
        return json.load(f)


# ---- スタイル適用 -----------------------------------------------------------

def apply_styles(ws, sheet_styles: dict) -> None:
    """
    sheet_styles（{ 座標: スタイルdict }）をワークシートに適用する。

    Font / PatternFill / Border / Alignment / number_format を復元。
    """
    for coord, style in sheet_styles.items():
        cell = ws[coord]

        # フォント
        if 'font' in style:
            fd = style['font']
            color = _make_color(fd.get('color'))
            cell.font = Font(
                name=fd.get('name', 'Calibri'),
                size=fd.get('size', 11.0),
                bold=fd.get('bold', False),
                italic=fd.get('italic', False),
                underline=fd.get('underline'),
                strike=fd.get('strike', False),
                color=color,
            )

        # 塗りつぶし
        if 'fill' in style:
            fd = style['fill']
            fg_color = _make_color(fd.get('fgColor'))
            bg_color = _make_color(fd.get('bgColor'))
            cell.fill = PatternFill(
                patternType=fd.get('patternType', 'solid'),
                fgColor=fg_color,
                bgColor=bg_color,
            )

        # 罫線
        if 'border' in style:
            bd = style['border']
            cell.border = Border(
                left=_make_side(bd.get('left')),
                right=_make_side(bd.get('right')),
                top=_make_side(bd.get('top')),
                bottom=_make_side(bd.get('bottom')),
                diagonal=_make_side(bd.get('diagonal')),
            )

        # 配置
        if 'alignment' in style:
            ad = style['alignment']
            cell.alignment = Alignment(
                horizontal=ad.get('horizontal'),
                vertical=ad.get('vertical'),
                wrap_text=ad.get('wrap_text', False),
                shrink_to_fit=ad.get('shrink_to_fit', False),
                indent=ad.get('indent', 0),
                text_rotation=ad.get('text_rotation', 0),
            )

        # 書式
        if 'number_format' in style:
            cell.number_format = style['number_format']


# ---- layout_map 構築 --------------------------------------------------------

def build_layout_map(outputs_dir: Path) -> dict[str, dict]:
    """
    layout.json を優先して読み込む。
    なければ旧 dimensions.json を互換フォーマットに変換して返す（後方互換）。
    どちらもなければ空 dict を返す。
    """
    layout_path = outputs_dir / 'layout.json'
    if layout_path.exists():
        with open(layout_path, encoding='utf-8') as f:
            return json.load(f)

    # 後方互換: dimensions.json → layout.json フォーマットに変換
    dimensions_path = outputs_dir / 'dimensions.json'
    if dimensions_path.exists():
        with open(dimensions_path, encoding='utf-8') as f:
            dimensions_data = json.load(f)

        layout_map: dict[str, dict] = {}
        for sheet_name, dims in dimensions_data.items():
            sheet_layout: dict = {}

            row_dims: dict[str, dict] = {}
            for row_num_str, height in dims.get('row_heights', {}).items():
                row_dims[row_num_str] = {'height': height}
            if row_dims:
                sheet_layout['row_dimensions'] = row_dims

            col_dims: dict[str, dict] = {}
            for col_letter, width in dims.get('col_widths', {}).items():
                col_dims[col_letter] = {'width': width}
            if col_dims:
                sheet_layout['col_dimensions'] = col_dims

            if sheet_layout:
                layout_map[sheet_name] = sheet_layout

        return layout_map

    return {}


# ---- レイアウト適用 ----------------------------------------------------------

def apply_layout(ws, sheet_layout: dict) -> None:
    """
    merged_cells / freeze_panes / row_dimensions / col_dimensions を復元。

    merged_cells:
        ws.merge_cells(range_string) を全て適用

    freeze_panes:
        ws.freeze_panes = sheet_layout["freeze_panes"]

    row_dimensions:
        ws.row_dimensions[n].height / hidden / customHeight を復元

    col_dimensions:
        ws.column_dimensions[c].width / hidden / customWidth を復元
    """
    # merged_cells
    for range_str in sheet_layout.get('merged_cells', []):
        ws.merge_cells(range_str)

    # freeze_panes
    fp = sheet_layout.get('freeze_panes')
    if fp:
        ws.freeze_panes = fp

    # row_dimensions
    for row_num_str, rd_info in sheet_layout.get('row_dimensions', {}).items():
        rd = ws.row_dimensions[int(row_num_str)]
        if 'height' in rd_info:
            rd.height = rd_info['height']
        if 'hidden' in rd_info:
            rd.hidden = rd_info['hidden']
        # customHeight は height 設定時に自動的に True になるため設定不要

    # col_dimensions
    for col_letter, cd_info in sheet_layout.get('col_dimensions', {}).items():
        cd = ws.column_dimensions[col_letter]
        if 'width' in cd_info:
            cd.width = cd_info['width']
        if 'hidden' in cd_info:
            cd.hidden = cd_info['hidden']
        # customWidth は width 設定時に自動的に True になるため設定不要


# ---- 条件付き書式復元ヘルパー ------------------------------------------------

def _make_dxf(d: dict) -> DifferentialStyle:
    """dict → DifferentialStyle。既存 _make_color / _make_side を再利用。"""
    kwargs: dict = {}

    if 'font' in d:
        fd = d['font']
        color = _make_color(fd.get('color'))
        kwargs['font'] = Font(
            bold=fd.get('bold', False),
            italic=fd.get('italic', False),
            underline=fd.get('underline'),
            strike=fd.get('strike', False),
            color=color,
        )

    if 'fill' in d:
        fd = d['fill']
        fg_color = _make_color(fd.get('fgColor'))
        bg_color = _make_color(fd.get('bgColor'))
        kwargs['fill'] = PatternFill(
            patternType=fd.get('patternType', 'solid'),
            fgColor=fg_color,
            bgColor=bg_color,
        )

    if 'border' in d:
        bd = d['border']
        kwargs['border'] = Border(
            left=_make_side(bd.get('left')),
            right=_make_side(bd.get('right')),
            top=_make_side(bd.get('top')),
            bottom=_make_side(bd.get('bottom')),
        )

    return DifferentialStyle(**kwargs)


def _make_cf_rule(rule_dict: dict) -> Rule:
    """dict → ConditionalFormattingRule"""
    rule_type = rule_dict.get('type', 'cellIs')
    priority = rule_dict.get('priority', 1)
    operator = rule_dict.get('operator')
    formula = rule_dict.get('formula', [])

    dxf_dict = rule_dict.get('dxf')
    dxf = _make_dxf(dxf_dict) if dxf_dict is not None else None

    # colorScale
    color_scale = None
    if rule_type == 'colorScale' and 'colorScale' in rule_dict:
        cs_data = rule_dict['colorScale']
        cfvo_list = [
            FormatObject(type=v.get('type', 'min'), val=v.get('val'))
            for v in cs_data.get('cfvo', [])
        ]
        color_list = [
            _make_color(c) for c in cs_data.get('color', [])
        ]
        color_scale = ColorScale(cfvo=cfvo_list, color=color_list)

    # dataBar
    data_bar = None
    if rule_type == 'dataBar' and 'dataBar' in rule_dict:
        db_data = rule_dict['dataBar']
        cfvo_list = [
            FormatObject(type=v.get('type', 'min'), val=v.get('val'))
            for v in db_data.get('cfvo', [])
        ]
        color_val = db_data.get('color')
        color = _make_color(color_val)
        data_bar = DataBar(cfvo=cfvo_list, color=color)

    # iconSet
    icon_set = None
    if rule_type == 'iconSet' and 'iconSet' in rule_dict:
        is_data = rule_dict['iconSet']
        cfvo_list = [
            FormatObject(type=v.get('type', 'percent'), val=v.get('val'))
            for v in is_data.get('cfvo', [])
        ]
        icon_set = IconSet(
            cfvo=cfvo_list,
            iconSet=is_data.get('iconSet', '3TrafficLights1'),
        )

    return Rule(
        type=rule_type,
        priority=priority,
        operator=operator,
        formula=formula,
        dxf=dxf,
        colorScale=color_scale,
        dataBar=data_bar,
        iconSet=icon_set,
    )


def apply_conditional_formatting(ws, cf_list: list) -> None:
    """
    conditional_formatting.json の1シート分を適用。
    ws.conditional_formatting.add(sqref, rule) を使用。
    """
    for cf_entry in cf_list:
        sqref = cf_entry.get('sqref', '')
        if not sqref:
            continue
        for rule_dict in cf_entry.get('rules', []):
            try:
                rule = _make_cf_rule(rule_dict)
                ws.conditional_formatting.add(sqref, rule)
            except Exception as e:
                print(f'    警告: 条件付き書式の復元をスキップ ({sqref}): {e}', file=sys.stderr)


# ---- 型変換 ------------------------------------------------------------------

def parse_value(s: str):
    """CSV 文字列 → Python 型変換"""
    if s == '':
        return None
    if s == 'TRUE':
        return True
    if s == 'FALSE':
        return False
    # 日時
    try:
        return datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        pass
    # 整数
    try:
        return int(s)
    except ValueError:
        pass
    # 浮動小数点
    try:
        return float(s)
    except ValueError:
        pass
    # 文字列
    return s


# ---- formula_map 構築 --------------------------------------------------------

def build_formula_map(cells_json: Path) -> dict[str, dict[str, dict]]:
    """
    cells.json を読み込み、{ シート名: { 座標: cell_info } } を返す。
    """
    if not cells_json.exists():
        return {}

    with open(cells_json, encoding='utf-8') as f:
        cells_data: dict[str, list] = json.load(f)

    formula_map: dict[str, dict[str, dict]] = {}
    for sheet_name, cells in cells_data.items():
        sheet_formulas: dict[str, dict] = {}
        for cell in cells:
            coord = cell.get('coord')
            if coord:
                sheet_formulas[coord] = cell
        formula_map[sheet_name] = sheet_formulas

    return formula_map


# ---- 1シート分の再構築 -------------------------------------------------------

def write_sheet(ws, csv_path: Path, sheet_formulas: dict) -> None:
    """
    CSV を読み込んでシートに書き込む。
    cells.json に座標が存在するセルは formula 文字列で上書きする。

    formula_type に関わらず formula 文字列をそのまま書き込む:
      - normal      : =len(P3) など → openpyxl が式として扱う
      - computed_value : IFERROR(__xludf... パターン → 式文字列として保持
      - importrange : IMPORTRANGE ヘッダー → 式文字列として保持
    """
    if not csv_path.exists():
        print(f'    警告: CSV が見つかりません: {csv_path}', file=sys.stderr)
        return

    with open(csv_path, encoding='utf-8', newline='') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, raw_value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                coord = cell.coordinate

                if coord in sheet_formulas:
                    # formula を書き込む（= で始まる文字列は openpyxl が式として扱う）
                    formula_str = sheet_formulas[coord].get('formula', '')
                    cell.value = formula_str
                else:
                    cell.value = parse_value(raw_value)


# ---- メイン処理 --------------------------------------------------------------

def reconstruct(outputs_dir: Path, output_path: Path) -> None:
    """
    outputs_dir/ の metadata.json・cells.json・styles.json・layout.json・
    conditional_formatting.json・csv/ から XLSX を再構築する。

    layout.json がない場合は旧 dimensions.json を読む（後方互換）。
    styles.json / layout.json / conditional_formatting.json が存在しない場合は
    それぞれなしで再構築（後方互換）。
    """
    meta_path = outputs_dir / 'metadata.json'
    cells_path = outputs_dir / 'cells.json'
    styles_path = outputs_dir / 'styles.json'
    cf_json_path = outputs_dir / 'conditional_formatting.json'

    if not meta_path.exists():
        print(f'エラー: metadata.json が見つかりません: {meta_path}', file=sys.stderr)
        sys.exit(1)

    print(f'再構築開始: {outputs_dir.name}')

    # [1] metadata.json 読み込み
    print('  [1/7] metadata.json 読み込み...')
    with open(meta_path, encoding='utf-8') as f:
        metadata = json.load(f)
    sheets = metadata.get('sheets', [])
    print(f'         {len(sheets)} シートを検出')

    # [2] cells.json → formula_map
    print('  [2/7] cells.json 読み込み...')
    formula_map = build_formula_map(cells_path)
    formula_sheet_count = len(formula_map)
    formula_cell_count = sum(len(v) for v in formula_map.values())
    print(f'         {formula_sheet_count} シート、{formula_cell_count:,} セルに式情報あり')

    # [3] styles.json → style_map（存在しない場合は空 dict）
    print('  [3/7] styles.json 読み込み...')
    style_map = build_style_map(styles_path)
    if style_map:
        style_cell_count = sum(len(v) for v in style_map.values())
        print(f'         {len(style_map)} シート、{style_cell_count:,} セルにスタイル情報あり')
    else:
        print('         styles.json なし（スタイルなしで再構築）')

    # [4] layout.json（or 旧 dimensions.json）→ layout_map
    print('  [4/7] layout.json 読み込み...')
    layout_map = build_layout_map(outputs_dir)
    if layout_map:
        layout_path = outputs_dir / 'layout.json'
        source = 'layout.json' if layout_path.exists() else 'dimensions.json（後方互換）'
        total_merged = sum(len(s.get('merged_cells', [])) for s in layout_map.values())
        sheets_with_freeze = sum(1 for s in layout_map.values() if s.get('freeze_panes'))
        print(f'         {source}: 結合セル {total_merged:,} 個、フリーズペイン {sheets_with_freeze} シート')
    else:
        print('         layout.json / dimensions.json なし（レイアウトなしで再構築）')

    # [5] conditional_formatting.json → cf_map（なければ空 dict）
    print('  [5/7] conditional_formatting.json 読み込み...')
    cf_map: dict[str, list] = {}
    if cf_json_path.exists():
        with open(cf_json_path, encoding='utf-8') as f:
            cf_map = json.load(f)
        cf_sheet_count = len(cf_map)
        cf_entry_count = sum(len(v) for v in cf_map.values())
        print(f'         {cf_sheet_count} シート、{cf_entry_count} エントリの条件付き書式あり')
    else:
        print('         conditional_formatting.json なし（条件付き書式なしで再構築）')

    # [6] ワークブック構築
    print('  [6/7] シート再構築中...')
    wb = openpyxl.Workbook()
    # デフォルトの "Sheet" を削除
    if wb.active:
        wb.remove(wb.active)

    for sheet_info in sheets:
        sheet_name = sheet_info['sheet_name']
        csv_rel_path = sheet_info.get('file', '')
        csv_path = outputs_dir / csv_rel_path

        ws = wb.create_sheet(title=sheet_name)
        sheet_formulas = formula_map.get(sheet_name, {})

        write_sheet(ws, csv_path, sheet_formulas)

        # スタイル適用
        sheet_styles = style_map.get(sheet_name, {})
        if sheet_styles:
            apply_styles(ws, sheet_styles)

        # レイアウト適用（結合セル・フリーズペイン・行高さ・列幅）
        sheet_layout = layout_map.get(sheet_name, {})
        if sheet_layout:
            apply_layout(ws, sheet_layout)

        # 条件付き書式適用
        sheet_cf = cf_map.get(sheet_name, [])
        if sheet_cf:
            apply_conditional_formatting(ws, sheet_cf)

        print(f'    ✓ {sheet_name}')

    # [7] 保存
    print(f'  [7/7] XLSX 保存中: {output_path}')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    size_kb = output_path.stat().st_size / 1024
    print(f'\n完了: {output_path}')
    print(f'  シート数      : {len(sheets)}')
    print(f'  ファイルサイズ : {size_kb:,.1f} KB')


# ---- CLI エントリーポイント --------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description='CSV/JSON → XLSX 再構築ツール（xlsx_to_csv_json.py の逆変換）'
    )
    parser.add_argument(
        'outputs_dir',
        type=Path,
        help='outputs/<name>/ ディレクトリのパス（metadata.json・cells.json・csv/ を含む）',
    )
    parser.add_argument(
        '--output', '-o',
        type=Path,
        default=None,
        help='出力 XLSX パス（省略時: outputs/<name>/<name>.xlsx）',
    )
    args = parser.parse_args()

    outputs_dir = args.outputs_dir.resolve()
    if not outputs_dir.is_dir():
        print(f'エラー: ディレクトリが見つかりません: {outputs_dir}', file=sys.stderr)
        sys.exit(1)

    if args.output:
        output_path = args.output.resolve()
    else:
        # デフォルト: outputs/<name>/<name>.xlsx
        output_path = outputs_dir / f'{outputs_dir.name}.xlsx'

    reconstruct(outputs_dir, output_path)


if __name__ == '__main__':
    main()
