#!/usr/bin/env python3
"""
コマ設計セクションにマスタデータ出力列を追加するスクリプト

「【ストーリー】必ず生きて帰る.xlsx」の各話シート（1話〜6話）の
コマ設計セクション右側に、MstPage / MstKomaLine 入力用のセル式列を追加する。

実際のXLSX行番号:
  - 行12: 基礎情報ヘッダー行（味方ゲートID, 敵ゲートID, ..., リリースキー）
  - 行13: 基礎情報データ行（N13 = リリースキー, P13 = ページID入力欄[新規追加]）
  - 行30: コマ設計ヘッダー行（B=行数, D=行パターンID, ...）
  - 行31〜35: コマ設計データ行（最大5行）

出力列（EA=131列〜FU=177列）を行30のヘッダー・行31〜35のデータ行に追加。

Usage:
    python3 scripts/add_koma_masterdata_output.py
    python3 scripts/add_koma_masterdata_output.py --output xlsx/カスタム出力.xlsx
"""

import sys
import shutil
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


# ---- 定数 -------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent
BASE_DIR = SCRIPT_DIR.parent

# 入力XLSXパス
SOURCE_XLSX = BASE_DIR / "xlsx" / "【ストーリー】必ず生きて帰る.xlsx"

# 出力XLSXパス（デフォルト）
DEFAULT_OUTPUT = BASE_DIR / "xlsx" / "【ストーリー】必ず生きて帰る_コマ設計出力追加.xlsx"

# 対象シート名
TARGET_SHEETS = ["1話", "2話", "3話", "4話", "5話", "6話"]

# 背景色（Enemyシーケンス右側出力と同じ薄い黄色）
HEADER_FILL_COLOR = "FFFFF2CC"
DATA_FILL_COLOR = "FFFFF2CC"

# 行番号（全シート共通）
BASIC_INFO_HEADER_ROW = 12   # 基礎情報ヘッダー行
BASIC_INFO_DATA_ROW = 13     # 基礎情報データ行（N13=リリースキー, P13=ページID）
KOMA_HEADER_ROW = 30         # コマ設計ヘッダー行
KOMA_DATA_START_ROW = 31     # コマ設計データ開始行
KOMA_DATA_END_ROW = 35       # コマ設計データ終了行

# 列番号（コマ設計セクション）
COL_D = 4    # 行パターンID
COL_H = 8    # コマ高さ
COL_J = 10   # コマ幅1
COL_L = 12   # コマ幅2
COL_N = 14   # コマ幅3（基礎情報ではリリースキー）
COL_P = 16   # コマ幅4（基礎情報ではページID入力欄）
COL_R = 18   # コマ背景
COL_U = 21   # コマ効果1
COL_Z = 26   # 効果時間（コマ効果1）
COL_AB = 28  # 効果数値（コマ効果1）

# 出力列（EA=131〜FU=177）
OUT_COL_EA = 131   # →MstPage ラベル
OUT_COL_EB = 132   # MstPage: id
OUT_COL_EC = 133   # MstPage: release_key
OUT_COL_EE = 135   # →MstKomaLine ラベル
OUT_COL_EF = 136   # MstKomaLine: id
OUT_COL_EG = 137   # MstKomaLine: mst_page_id
OUT_COL_EH = 138   # MstKomaLine: row
OUT_COL_EI = 139   # MstKomaLine: height
OUT_COL_EJ = 140   # MstKomaLine: koma_line_layout_asset_key
OUT_COL_EK = 141   # MstKomaLine: koma1_asset_key
OUT_COL_EL = 142   # MstKomaLine: koma1_width
OUT_COL_EM = 143   # MstKomaLine: koma1_back_ground_offset
OUT_COL_EN = 144   # MstKomaLine: koma1_effect_type
OUT_COL_EO = 145   # MstKomaLine: koma1_effect_parameter1
OUT_COL_EP = 146   # MstKomaLine: koma1_effect_parameter2
OUT_COL_EQ = 147   # MstKomaLine: koma1_effect_target_side
OUT_COL_ER = 148   # MstKomaLine: koma1_effect_target_colors
OUT_COL_ES = 149   # MstKomaLine: koma1_effect_target_roles
OUT_COL_ET = 150   # MstKomaLine: koma2_asset_key
OUT_COL_EU = 151   # MstKomaLine: koma2_width
OUT_COL_EV = 152   # MstKomaLine: koma2_back_ground_offset
OUT_COL_EW = 153   # MstKomaLine: koma2_effect_type
OUT_COL_EX = 154   # MstKomaLine: koma2_effect_parameter1
OUT_COL_EY = 155   # MstKomaLine: koma2_effect_parameter2
OUT_COL_EZ = 156   # MstKomaLine: koma2_effect_target_side
OUT_COL_FA = 157   # MstKomaLine: koma2_effect_target_colors
OUT_COL_FB = 158   # MstKomaLine: koma2_effect_target_roles
OUT_COL_FC = 159   # MstKomaLine: koma3_asset_key
OUT_COL_FD = 160   # MstKomaLine: koma3_width
OUT_COL_FE = 161   # MstKomaLine: koma3_back_ground_offset
OUT_COL_FF = 162   # MstKomaLine: koma3_effect_type
OUT_COL_FG = 163   # MstKomaLine: koma3_effect_parameter1
OUT_COL_FH = 164   # MstKomaLine: koma3_effect_parameter2
OUT_COL_FI = 165   # MstKomaLine: koma3_effect_target_side
OUT_COL_FJ = 166   # MstKomaLine: koma3_effect_target_colors
OUT_COL_FK = 167   # MstKomaLine: koma3_effect_target_roles
OUT_COL_FL = 168   # MstKomaLine: koma4_asset_key
OUT_COL_FM = 169   # MstKomaLine: koma4_width
OUT_COL_FN = 170   # MstKomaLine: koma4_back_ground_offset
OUT_COL_FO = 171   # MstKomaLine: koma4_effect_type
OUT_COL_FP = 172   # MstKomaLine: koma4_effect_parameter1
OUT_COL_FQ = 173   # MstKomaLine: koma4_effect_parameter2
OUT_COL_FR = 174   # MstKomaLine: koma4_effect_target_side
OUT_COL_FS = 175   # MstKomaLine: koma4_effect_target_colors
OUT_COL_FT = 176   # MstKomaLine: koma4_effect_target_roles
OUT_COL_FU = 177   # MstKomaLine: release_key


# ---- ヘルパー -------------------------------------------------------------------

def make_fill(color: str) -> PatternFill:
    """背景色塗りつぶしを作成"""
    return PatternFill(patternType="solid", fgColor=color)


def set_cell(ws, row: int, col: int, value, fill_color: str | None = None) -> None:
    """セルに値（または式）を設定し、オプションで背景色を適用"""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if fill_color:
        cell.fill = make_fill(fill_color)


# ---- 基礎情報セクション処理 ---------------------------------------------------

def add_page_id_field(ws) -> None:
    """
    基礎情報セクション（行12〜13）にページID入力欄を追加する。

    現状: N12:P12 と N13:P13 が結合セルになっている
    変更: 結合を解除して P12/P13 をページID用に使用する

    - P12: ヘッダー「ページID」
    - P13: データ入力セル（ユーザーが入力する）
    """
    # 既存の結合 N12:P12 を解除する
    try:
        ws.unmerge_cells("N12:P12")
    except Exception:
        pass  # 既に結合がない場合はスキップ

    # 既存の結合 N13:P13 を解除する
    try:
        ws.unmerge_cells("N13:P13")
    except Exception:
        pass

    # P12 にページIDヘッダーを設定
    ws.cell(row=BASIC_INFO_HEADER_ROW, column=COL_P).value = "ページID"

    # P13 はユーザー入力欄として空白のまま（既存値があれば上書きしない）
    existing = ws.cell(row=BASIC_INFO_DATA_ROW, column=COL_P).value
    if existing is None:
        ws.cell(row=BASIC_INFO_DATA_ROW, column=COL_P).value = None


# ---- コマ設計ヘッダー行 -------------------------------------------------------

HEADER_LABELS = {
    OUT_COL_EA: "→MstPage",
    OUT_COL_EB: "id",
    OUT_COL_EC: "release_key",
    # OUT_COL_ED: (空)
    OUT_COL_EE: "→MstKomaLine",
    OUT_COL_EF: "id",
    OUT_COL_EG: "mst_page_id",
    OUT_COL_EH: "row",
    OUT_COL_EI: "height",
    OUT_COL_EJ: "koma_line_layout_asset_key",
    OUT_COL_EK: "koma1_asset_key",
    OUT_COL_EL: "koma1_width",
    OUT_COL_EM: "koma1_back_ground_offset",
    OUT_COL_EN: "koma1_effect_type",
    OUT_COL_EO: "koma1_effect_parameter1",
    OUT_COL_EP: "koma1_effect_parameter2",
    OUT_COL_EQ: "koma1_effect_target_side",
    OUT_COL_ER: "koma1_effect_target_colors",
    OUT_COL_ES: "koma1_effect_target_roles",
    OUT_COL_ET: "koma2_asset_key",
    OUT_COL_EU: "koma2_width",
    OUT_COL_EV: "koma2_back_ground_offset",
    OUT_COL_EW: "koma2_effect_type",
    OUT_COL_EX: "koma2_effect_parameter1",
    OUT_COL_EY: "koma2_effect_parameter2",
    OUT_COL_EZ: "koma2_effect_target_side",
    OUT_COL_FA: "koma2_effect_target_colors",
    OUT_COL_FB: "koma2_effect_target_roles",
    OUT_COL_FC: "koma3_asset_key",
    OUT_COL_FD: "koma3_width",
    OUT_COL_FE: "koma3_back_ground_offset",
    OUT_COL_FF: "koma3_effect_type",
    OUT_COL_FG: "koma3_effect_parameter1",
    OUT_COL_FH: "koma3_effect_parameter2",
    OUT_COL_FI: "koma3_effect_target_side",
    OUT_COL_FJ: "koma3_effect_target_colors",
    OUT_COL_FK: "koma3_effect_target_roles",
    OUT_COL_FL: "koma4_asset_key",
    OUT_COL_FM: "koma4_width",
    OUT_COL_FN: "koma4_back_ground_offset",
    OUT_COL_FO: "koma4_effect_type",
    OUT_COL_FP: "koma4_effect_parameter1",
    OUT_COL_FQ: "koma4_effect_parameter2",
    OUT_COL_FR: "koma4_effect_target_side",
    OUT_COL_FS: "koma4_effect_target_colors",
    OUT_COL_FT: "koma4_effect_target_roles",
    OUT_COL_FU: "release_key",
}


def add_koma_header(ws) -> None:
    """
    コマ設計ヘッダー行（行30）のEA列以降にMstPage/MstKomaLineヘッダーを追加する。
    """
    for col, label in HEADER_LABELS.items():
        set_cell(ws, KOMA_HEADER_ROW, col, label, HEADER_FILL_COLOR)


# ---- コマ設計データ行 ---------------------------------------------------------

def add_koma_data_formulas(ws) -> None:
    """
    コマ設計データ行（行31〜35）にMstPage/MstKomaLine出力用のセル式を追加する。

    - MstPage: id は行31のみ（ページ全体で1レコード）
    - MstKomaLine: 各行（行31〜35）にそれぞれのレコード

    ページIDは $P$13 を参照（ユーザーが基礎情報セクションに入力した値）
    リリースキーは $N$13 を参照
    """
    # -- MstPage（行31のみに追加）--
    # id: ページID入力欄を参照
    set_cell(ws, KOMA_DATA_START_ROW, OUT_COL_EB, "=$P$13", DATA_FILL_COLOR)
    # release_key: リリースキー参照
    set_cell(ws, KOMA_DATA_START_ROW, OUT_COL_EC, "=$N$13", DATA_FILL_COLOR)

    # -- MstKomaLine（行31〜35）--
    for r in range(KOMA_DATA_START_ROW, KOMA_DATA_END_ROW + 1):
        row_offset = r - KOMA_HEADER_ROW  # 1〜5

        # 空行ガード条件（H列=コマ高さが空なら全列空白）
        h_empty = f'H{r}=""'

        def formula(expr: str) -> str:
            """空行ガード付き式を生成"""
            return f'=IF({h_empty},"",{expr})'

        def koma_null_formula(width_col: str, expr: str) -> str:
            """コマ幅が"none"なら__NULL__、それ以外は expr"""
            return f'=IF({h_empty},"",IF({width_col}{r}="none","__NULL__",{expr}))'

        # EF: id
        set_cell(ws, r, OUT_COL_EF,
                 formula(f'CONCATENATE("koma_line_",$P$13,"_{row_offset}")'),
                 DATA_FILL_COLOR)

        # EG: mst_page_id
        set_cell(ws, r, OUT_COL_EG, formula("$P$13"), DATA_FILL_COLOR)

        # EH: row
        set_cell(ws, r, OUT_COL_EH, formula(str(row_offset)), DATA_FILL_COLOR)

        # EI: height
        set_cell(ws, r, OUT_COL_EI, formula(f"H{r}"), DATA_FILL_COLOR)

        # EJ: koma_line_layout_asset_key
        set_cell(ws, r, OUT_COL_EJ, formula(f"D{r}"), DATA_FILL_COLOR)

        # EK: koma1_asset_key（コマ背景=R列）
        set_cell(ws, r, OUT_COL_EK, formula(f"R{r}"), DATA_FILL_COLOR)

        # EL: koma1_width
        set_cell(ws, r, OUT_COL_EL, formula(f"J{r}"), DATA_FILL_COLOR)

        # EM: koma1_back_ground_offset（固定値 -1）
        set_cell(ws, r, OUT_COL_EM, formula('"-1"'), DATA_FILL_COLOR)

        # EN: koma1_effect_type
        set_cell(ws, r, OUT_COL_EN,
                 formula(f'IF(U{r}="","None",U{r})'),
                 DATA_FILL_COLOR)

        # EO: koma1_effect_parameter1（効果時間）
        set_cell(ws, r, OUT_COL_EO,
                 formula(f'IF(Z{r}="","0",Z{r})'),
                 DATA_FILL_COLOR)

        # EP: koma1_effect_parameter2（効果数値）
        set_cell(ws, r, OUT_COL_EP,
                 formula(f'IF(AB{r}="","0",AB{r})'),
                 DATA_FILL_COLOR)

        # EQ: koma1_effect_target_side（固定値 "All"）
        set_cell(ws, r, OUT_COL_EQ, formula('"All"'), DATA_FILL_COLOR)

        # ER: koma1_effect_target_colors（固定値 "All"）
        set_cell(ws, r, OUT_COL_ER, formula('"All"'), DATA_FILL_COLOR)

        # ES: koma1_effect_target_roles（固定値 "All"）
        set_cell(ws, r, OUT_COL_ES, formula('"All"'), DATA_FILL_COLOR)

        # -- koma2（L列=コマ幅2が"none"なら__NULL__）--
        # ET: koma2_asset_key
        set_cell(ws, r, OUT_COL_ET, koma_null_formula("L", f"R{r}"), DATA_FILL_COLOR)
        # EU: koma2_width
        set_cell(ws, r, OUT_COL_EU, koma_null_formula("L", f"L{r}"), DATA_FILL_COLOR)
        # EV: koma2_back_ground_offset
        set_cell(ws, r, OUT_COL_EV, koma_null_formula("L", '"-1"'), DATA_FILL_COLOR)
        # EW: koma2_effect_type
        set_cell(ws, r, OUT_COL_EW, koma_null_formula("L", '"None"'), DATA_FILL_COLOR)
        # EX: koma2_effect_parameter1
        set_cell(ws, r, OUT_COL_EX, koma_null_formula("L", '"0"'), DATA_FILL_COLOR)
        # EY: koma2_effect_parameter2
        set_cell(ws, r, OUT_COL_EY, koma_null_formula("L", '"0"'), DATA_FILL_COLOR)
        # EZ: koma2_effect_target_side
        set_cell(ws, r, OUT_COL_EZ, koma_null_formula("L", '"All"'), DATA_FILL_COLOR)
        # FA: koma2_effect_target_colors
        set_cell(ws, r, OUT_COL_FA, koma_null_formula("L", '"All"'), DATA_FILL_COLOR)
        # FB: koma2_effect_target_roles
        set_cell(ws, r, OUT_COL_FB, koma_null_formula("L", '"All"'), DATA_FILL_COLOR)

        # -- koma3（N列=コマ幅3が"none"なら__NULL__）--
        # FC: koma3_asset_key
        set_cell(ws, r, OUT_COL_FC, koma_null_formula("N", f"R{r}"), DATA_FILL_COLOR)
        # FD: koma3_width
        set_cell(ws, r, OUT_COL_FD, koma_null_formula("N", f"N{r}"), DATA_FILL_COLOR)
        # FE: koma3_back_ground_offset
        set_cell(ws, r, OUT_COL_FE, koma_null_formula("N", '"-1"'), DATA_FILL_COLOR)
        # FF: koma3_effect_type
        set_cell(ws, r, OUT_COL_FF, koma_null_formula("N", '"None"'), DATA_FILL_COLOR)
        # FG: koma3_effect_parameter1
        set_cell(ws, r, OUT_COL_FG, koma_null_formula("N", '"0"'), DATA_FILL_COLOR)
        # FH: koma3_effect_parameter2
        set_cell(ws, r, OUT_COL_FH, koma_null_formula("N", '"0"'), DATA_FILL_COLOR)
        # FI: koma3_effect_target_side
        set_cell(ws, r, OUT_COL_FI, koma_null_formula("N", '"All"'), DATA_FILL_COLOR)
        # FJ: koma3_effect_target_colors
        set_cell(ws, r, OUT_COL_FJ, koma_null_formula("N", '"All"'), DATA_FILL_COLOR)
        # FK: koma3_effect_target_roles
        set_cell(ws, r, OUT_COL_FK, koma_null_formula("N", '"All"'), DATA_FILL_COLOR)

        # -- koma4（P列=コマ幅4が"none"なら__NULL__）--
        # FL: koma4_asset_key
        set_cell(ws, r, OUT_COL_FL, koma_null_formula("P", f"R{r}"), DATA_FILL_COLOR)
        # FM: koma4_width
        set_cell(ws, r, OUT_COL_FM, koma_null_formula("P", f"P{r}"), DATA_FILL_COLOR)
        # FN: koma4_back_ground_offset
        set_cell(ws, r, OUT_COL_FN, koma_null_formula("P", '"-1"'), DATA_FILL_COLOR)
        # FO: koma4_effect_type
        set_cell(ws, r, OUT_COL_FO, koma_null_formula("P", '"None"'), DATA_FILL_COLOR)
        # FP: koma4_effect_parameter1
        set_cell(ws, r, OUT_COL_FP, koma_null_formula("P", '"0"'), DATA_FILL_COLOR)
        # FQ: koma4_effect_parameter2
        set_cell(ws, r, OUT_COL_FQ, koma_null_formula("P", '"0"'), DATA_FILL_COLOR)
        # FR: koma4_effect_target_side
        set_cell(ws, r, OUT_COL_FR, koma_null_formula("P", '"All"'), DATA_FILL_COLOR)
        # FS: koma4_effect_target_colors
        set_cell(ws, r, OUT_COL_FS, koma_null_formula("P", '"All"'), DATA_FILL_COLOR)
        # FT: koma4_effect_target_roles
        set_cell(ws, r, OUT_COL_FT, koma_null_formula("P", '"All"'), DATA_FILL_COLOR)

        # FU: release_key
        set_cell(ws, r, OUT_COL_FU, formula("$N$13"), DATA_FILL_COLOR)

    # -- MstPage の EA列（ラベル行31に →MstPage のマーカーは行30のみ。データ行ではid/release_keyのみ）
    # EA列（OUT_COL_EA=131）はヘッダー行30のみラベルあり。データ行はEA列を使用しない。
    # （MstPageはEB/EC列がid/release_keyに対応）


# ---- シート処理 ---------------------------------------------------------------

def process_sheet(ws, sheet_name: str) -> None:
    """1シート分の処理（基礎情報追加 + コマ設計右側出力追加）"""
    print(f"  処理中: {sheet_name}")

    # 1. 基礎情報セクションにページID入力欄を追加
    add_page_id_field(ws)

    # 2. コマ設計ヘッダー行にMstPage/MstKomaLineヘッダーを追加
    add_koma_header(ws)

    # 3. コマ設計データ行にセル式を追加
    add_koma_data_formulas(ws)

    print(f"    ✓ ページID入力欄追加（P{BASIC_INFO_HEADER_ROW}〜P{BASIC_INFO_DATA_ROW}）")
    print(f"    ✓ MstPageヘッダー追加（行{KOMA_HEADER_ROW}, EB〜EC列）")
    print(f"    ✓ MstKomaLineヘッダー追加（行{KOMA_HEADER_ROW}, EE〜FU列）")
    print(f"    ✓ コマ設計右側出力追加（行{KOMA_DATA_START_ROW}〜{KOMA_DATA_END_ROW}, EB〜FU列）")


# ---- メイン ------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="コマ設計セクションにマスタデータ出力列を追加する"
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        help=f"出力XLSXパス（デフォルト: {DEFAULT_OUTPUT}）",
    )
    parser.add_argument(
        "--source", "-s",
        type=Path,
        default=None,
        help=f"入力XLSXパス（デフォルト: {SOURCE_XLSX}）",
    )
    args = parser.parse_args()

    source_path = (args.source or SOURCE_XLSX).resolve()
    output_path = (args.output or DEFAULT_OUTPUT).resolve()

    # 入力ファイルの存在確認
    if not source_path.exists():
        print(f"エラー: 入力XLSXが見つかりません: {source_path}", file=sys.stderr)
        sys.exit(1)

    print(f"入力: {source_path}")
    print(f"出力: {output_path}")

    # [1] XLSXを読み込む（data_only=False で式を保持）
    print("\n[1/3] XLSXを読み込み中...")
    wb = openpyxl.load_workbook(str(source_path), data_only=False)

    # [2] 各話シートを処理
    print("[2/3] シート処理中...")
    processed = 0
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  警告: シート '{sheet_name}' が見つかりません。スキップします。",
                  file=sys.stderr)
            continue
        ws = wb[sheet_name]
        process_sheet(ws, sheet_name)
        processed += 1

    # [3] 保存
    print(f"\n[3/3] XLSX保存中: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    size_kb = output_path.stat().st_size / 1024
    print(f"\n完了!")
    print(f"  処理シート数: {processed}")
    print(f"  ファイルサイズ: {size_kb:,.1f} KB")
    print(f"  出力先: {output_path}")
    print()
    print("次のステップ:")
    print("  1. 各話シートのP13セル（ページID欄）に適切なページIDを入力してください")
    print("     例: page_jig1_00101（1話）, page_jig1_00102（2話）等")
    print("  2. コマ設計行（行31〜35）の右側出力（EB〜FU列）で計算結果を確認してください")


if __name__ == "__main__":
    main()
