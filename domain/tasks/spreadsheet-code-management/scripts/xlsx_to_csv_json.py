#!/usr/bin/env python3
"""
XLSX → CSV / JSON 変換スクリプト

Google Spreadsheet を XLSX でダウンロードし、コード管理できる形式に変換する。

出力:
  <output_dir>/
    csv/
      <シート名>.csv    # 実データ（計算済み値）
    cells.json          # 式情報（どのセルにどんな式があるか）
    metadata.json       # ブック全体のメタ情報
    styles.json         # セルスタイル情報（色・フォント・罫線・配置・書式）
    layout.json         # レイアウト情報（結合セル・フリーズペイン・行高さ・列幅）
    conditional_formatting.json  # 条件付き書式

Usage:
    python3 scripts/xlsx_to_csv_json.py <xlsx_file>
    python3 scripts/xlsx_to_csv_json.py <xlsx_file> --output-dir outputs/my_sheet
"""

import re
import sys
import csv
import json
import argparse
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


# ---- 定数 / 正規表現 -------------------------------------------------------

# IMPORTRANGE ヘッダーパターン:
#   =IFERROR(__xludf.DUMMYFUNCTION("IMPORTRANGE(""url"",""range"")"), "")
_RE_IMPORTRANGE = re.compile(
    r'=IFERROR\(__xludf\.DUMMYFUNCTION\("IMPORTRANGE\(""(.+?)"",\s*""(.+?)""\)"\)',
    re.DOTALL,
)

# COMPUTED_VALUE パターン（IMPORTRANGE で転記されたデータセル）:
#   =IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE"""), "actual_value")
_COMPUTED_VALUE_PREFIX = '=IFERROR(__xludf.DUMMYFUNCTION("""COMPUTED_VALUE""")'


# ---- ユーティリティ ---------------------------------------------------------

def sanitize_filename(name: str) -> str:
    """シート名をファイルシステムで安全な文字列に変換"""
    for ch in r'\/:*?"<>|▶︎':
        name = name.replace(ch, '_')
    return name.strip() or 'sheet'


def cell_value_to_str(value) -> str:
    """セル値を CSV 用文字列に変換"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else str(value)
    return str(value)


def extract_iferror_fallback(formula: str) -> str | None:
    """
    IFERROR(expr, fallback) の fallback 部分を抽出する。
    ネストされた括弧・文字列リテラルを考慮した簡易パーサ。
    """
    # "=IFERROR(" の後ろから深さ1の最後のカンマを探す
    depth = 0
    in_str = False
    last_comma = -1
    i = 0
    while i < len(formula):
        ch = formula[i]
        if ch == '"':
            # 連続する "" は単一のリテラル " として扱う
            if in_str and i + 1 < len(formula) and formula[i + 1] == '"':
                i += 2
                continue
            in_str = not in_str
        elif not in_str:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif ch == ',' and depth == 1:
                last_comma = i
        i += 1

    if last_comma == -1:
        return None

    # fallback 部分: カンマの後〜末尾の ')' の前
    fallback = formula[last_comma + 1:].strip()
    if fallback.endswith(')'):
        fallback = fallback[:-1].strip()

    # 前後のダブルクォートを除去
    if len(fallback) >= 2 and fallback[0] == '"' and fallback[-1] == '"':
        # "" のエスケープを戻す
        return fallback[1:-1].replace('""', '"')

    return fallback if fallback else None


def classify_formula(formula: str) -> dict:
    """
    式を分類してメタ情報を返す。

    returns: {
        "formula_type": "importrange" | "computed_value" | "normal",
        ...追加情報...
    }
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return {"formula_type": "normal", "formula": formula}

    # IMPORTRANGE ヘッダー
    m = _RE_IMPORTRANGE.match(formula)
    if m:
        return {
            "formula_type": "importrange",
            "formula": formula,
            "importrange_url": m.group(1),
            "importrange_range": m.group(2),
        }

    # COMPUTED_VALUE（IMPORTRANGE データセル）
    if formula.startswith(_COMPUTED_VALUE_PREFIX):
        fallback = extract_iferror_fallback(formula)
        return {
            "formula_type": "computed_value",
            "formula": formula,
            "value": fallback,
        }

    # 通常の式（=len(G3), =IFERROR(...) など）
    return {
        "formula_type": "normal",
        "formula": formula,
    }


# ---- スタイル抽出ヘルパー ---------------------------------------------------

def _extract_color(color) -> str | dict | None:
    """
    openpyxl の Color オブジェクトを JSON シリアライズ可能な値に変換。

    - rgb 型 → "RRGGBBAA" 文字列
    - theme=1, tint=0.0 → None（デフォルト黒はスキップ）
    - theme 型（その他） → {"theme": N, "tint": T}
    - indexed 型 → {"indexed": N}
    """
    if color is None:
        return None

    color_type = getattr(color, 'type', None)

    if color_type == 'rgb':
        rgb = color.rgb
        if rgb in ('00000000', '000000', None):
            return None
        return rgb

    if color_type == 'theme':
        theme = getattr(color, 'theme', None)
        tint = getattr(color, 'tint', 0.0)
        # theme=1, tint=0.0 はデフォルト黒なのでスキップ
        if theme == 1 and tint == 0.0:
            return None
        return {'theme': theme, 'tint': tint}

    if color_type == 'indexed':
        indexed = getattr(color, 'indexed', None)
        return {'indexed': indexed}

    return None


def _extract_color_raw(color) -> str | dict | None:
    """
    openpyxl の Color オブジェクトを省略なしで変換（条件付き書式など用）。
    _extract_color と異なり '00000000' も保持する。
    """
    if color is None:
        return None

    color_type = getattr(color, 'type', None)

    if color_type == 'rgb':
        rgb = getattr(color, 'rgb', None)
        return rgb if rgb else None

    if color_type == 'theme':
        theme = getattr(color, 'theme', None)
        tint = getattr(color, 'tint', 0.0)
        return {'theme': theme, 'tint': tint}

    if color_type == 'indexed':
        indexed = getattr(color, 'indexed', None)
        return {'indexed': indexed}

    return None


def _extract_font(font) -> dict:
    """フォント情報を抽出。デフォルト値はスキップ。"""
    result = {}
    if font is None:
        return result

    name = getattr(font, 'name', None)
    if name and name != 'Calibri':
        result['name'] = name

    size = getattr(font, 'size', None)
    if size and size != 11.0:
        result['size'] = size

    bold = getattr(font, 'bold', None)
    if bold:
        result['bold'] = bold

    italic = getattr(font, 'italic', None)
    if italic:
        result['italic'] = italic

    underline = getattr(font, 'underline', None)
    if underline:
        result['underline'] = underline

    strike = getattr(font, 'strike', None)
    if strike:
        result['strike'] = strike

    color = _extract_color(getattr(font, 'color', None))
    if color is not None:
        result['color'] = color

    return result


def _extract_fill(fill) -> dict:
    """
    塗りつぶし情報を抽出。

    patternType='none' はスキップ。
    fgColor が '00000000' でも patternType='solid' なら保存（白背景として有効）。
    """
    result = {}
    if fill is None:
        return result

    pattern_type = getattr(fill, 'patternType', None)
    if not pattern_type or pattern_type == 'none':
        return result

    result['patternType'] = pattern_type

    fg = _extract_color(getattr(fill, 'fgColor', None))
    # fgColor='00000000' かつ patternType='solid' → 黒背景として保存
    # ただし color type='rgb' で値が '00000000' の場合は _extract_color が None を返すので
    # 直接チェックして保存する
    fg_raw = getattr(fill, 'fgColor', None)
    if fg_raw is not None and getattr(fg_raw, 'type', None) == 'rgb':
        rgb_val = fg_raw.rgb
        if rgb_val == '00000000' and pattern_type == 'solid':
            result['fgColor'] = rgb_val
        elif fg is not None:
            result['fgColor'] = fg
    elif fg is not None:
        result['fgColor'] = fg

    bg = _extract_color(getattr(fill, 'bgColor', None))
    if bg is not None:
        result['bgColor'] = bg

    return result


def _extract_border_side(side) -> dict | None:
    """罫線の1辺を抽出。side が None の場合は None を返す。"""
    if side is None:
        return None

    border_style = getattr(side, 'border_style', None)
    if not border_style:
        return None

    result: dict = {'border_style': border_style}
    color = _extract_color(getattr(side, 'color', None))
    if color is not None:
        result['color'] = color

    return result


def _extract_alignment(al) -> dict:
    """配置情報を抽出。デフォルト値はスキップ。"""
    result = {}
    if al is None:
        return result

    horizontal = getattr(al, 'horizontal', None)
    if horizontal:
        result['horizontal'] = horizontal

    vertical = getattr(al, 'vertical', None)
    if vertical:
        result['vertical'] = vertical

    wrap_text = getattr(al, 'wrap_text', None)
    if wrap_text:
        result['wrap_text'] = wrap_text

    shrink_to_fit = getattr(al, 'shrink_to_fit', None)
    if shrink_to_fit:
        result['shrink_to_fit'] = shrink_to_fit

    indent = getattr(al, 'indent', None)
    if indent:
        result['indent'] = indent

    text_rotation = getattr(al, 'text_rotation', None)
    if text_rotation:
        result['text_rotation'] = text_rotation

    return result


# ---- styles.json 出力 -------------------------------------------------------

def export_styles_json(wb_data: openpyxl.Workbook, output_dir: Path) -> Path:
    """
    全シートの全セルのスタイル情報を styles.json に出力する。

    スタイルが何もないセルは JSON に含めない（省スペース）。

    出力形式:
    {
      "シート名": {
        "A1": {
          "font": {...},
          "fill": {...},
          "border": {"left": {...}, "right": {...}, ...},
          "alignment": {...},
          "number_format": "..."
        },
        ...
      },
      ...
    }
    """
    styles_data: dict[str, dict] = {}

    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]
        sheet_styles: dict[str, dict] = {}

        for row in ws.iter_rows():
            for cell in row:
                cell_style: dict = {}

                # フォント
                font = _extract_font(getattr(cell, 'font', None))
                if font:
                    cell_style['font'] = font

                # 塗りつぶし
                fill = _extract_fill(getattr(cell, 'fill', None))
                if fill:
                    cell_style['fill'] = fill

                # 罫線
                border_obj = getattr(cell, 'border', None)
                if border_obj:
                    border: dict = {}
                    for side_name in ('left', 'right', 'top', 'bottom', 'diagonal'):
                        side = _extract_border_side(getattr(border_obj, side_name, None))
                        if side:
                            border[side_name] = side
                    if border:
                        cell_style['border'] = border

                # 配置
                alignment = _extract_alignment(getattr(cell, 'alignment', None))
                if alignment:
                    cell_style['alignment'] = alignment

                # 書式（デフォルト 'General' はスキップ）
                number_format = getattr(cell, 'number_format', None)
                if number_format and number_format != 'General':
                    cell_style['number_format'] = number_format

                if cell_style:
                    sheet_styles[cell.coordinate] = cell_style

        if sheet_styles:
            styles_data[sheet_name] = sheet_styles

    styles_path = output_dir / 'styles.json'
    with open(styles_path, 'w', encoding='utf-8') as f:
        json.dump(styles_data, f, ensure_ascii=False, indent=2)

    return styles_path


# ---- layout.json 出力 -------------------------------------------------------

def export_layout_json(wb_data: openpyxl.Workbook, output_dir: Path) -> Path:
    """
    全シートのレイアウト情報を layout.json に出力。

    - merged_cells: str(range) のリスト
    - freeze_panes: セル座標文字列 or None
    - row_dimensions: height / hidden / customHeight（全てデフォルト値の行はスキップ）
    - col_dimensions: width / hidden / customWidth（全てデフォルト値の列はスキップ）

    出力形式:
    {
      "シート名": {
        "merged_cells": ["A1:C3", "D5:F5"],
        "freeze_panes": "A2",
        "row_dimensions": {
          "1": {"height": 20.0, "hidden": false, "customHeight": true}
        },
        "col_dimensions": {
          "A": {"width": 21.0, "hidden": false, "customWidth": true}
        }
      }
    }
    """
    layout_data: dict[str, dict] = {}

    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]
        sheet_layout: dict = {}

        # merged_cells
        merged = [str(m) for m in ws.merged_cells.ranges]
        if merged:
            sheet_layout['merged_cells'] = sorted(merged)

        # freeze_panes
        fp = ws.freeze_panes
        if fp:
            sheet_layout['freeze_panes'] = str(fp)

        # row_dimensions
        row_dims: dict[str, dict] = {}
        for row_num, rd in ws.row_dimensions.items():
            entry: dict = {}
            if rd.height is not None:
                entry['height'] = rd.height
                entry['customHeight'] = bool(getattr(rd, 'customHeight', True))
            if getattr(rd, 'hidden', False):
                entry['hidden'] = True
            if entry:
                row_dims[str(row_num)] = entry
        if row_dims:
            sheet_layout['row_dimensions'] = row_dims

        # col_dimensions
        col_dims: dict[str, dict] = {}
        for col_letter, cd in ws.column_dimensions.items():
            entry = {}
            if cd.width is not None:
                entry['width'] = cd.width
                entry['customWidth'] = bool(getattr(cd, 'customWidth', True))
            if getattr(cd, 'hidden', False):
                entry['hidden'] = True
            if entry:
                col_dims[col_letter] = entry
        if col_dims:
            sheet_layout['col_dimensions'] = col_dims

        if sheet_layout:
            layout_data[sheet_name] = sheet_layout

    layout_path = output_dir / 'layout.json'
    with open(layout_path, 'w', encoding='utf-8') as f:
        json.dump(layout_data, f, ensure_ascii=False, indent=2)

    return layout_path


# ---- conditional_formatting.json 出力 ---------------------------------------

def _serialize_dxf(dxf) -> dict:
    """
    DifferentialStyle → dict。
    既存の _extract_font / _extract_fill / _extract_border_side を再利用。
    """
    result: dict = {}
    if dxf is None:
        return result

    font = _extract_font(getattr(dxf, 'font', None))
    if font:
        result['font'] = font

    fill = _extract_fill(getattr(dxf, 'fill', None))
    if fill:
        result['fill'] = fill

    border_obj = getattr(dxf, 'border', None)
    if border_obj:
        border: dict = {}
        for side_name in ('left', 'right', 'top', 'bottom'):
            side = _extract_border_side(getattr(border_obj, side_name, None))
            if side:
                border[side_name] = side
        if border:
            result['border'] = border

    return result


def _serialize_cf_rule(rule) -> dict:
    """
    ConditionalFormattingRule → dict。
    type / operator / formula / priority / dxf / colorScale / dataBar / iconSet を含む。
    """
    result: dict = {}

    rule_type = getattr(rule, 'type', None)
    if rule_type:
        result['type'] = rule_type

    operator = getattr(rule, 'operator', None)
    if operator:
        result['operator'] = operator

    formula = getattr(rule, 'formula', None)
    if formula:
        result['formula'] = list(formula)

    priority = getattr(rule, 'priority', None)
    if priority is not None:
        result['priority'] = priority

    dxf = getattr(rule, 'dxf', None)
    if dxf is not None:
        result['dxf'] = _serialize_dxf(dxf)

    # colorScale
    color_scale = getattr(rule, 'colorScale', None)
    if color_scale is not None:
        cs_data: dict = {}
        cfvo_list = getattr(color_scale, 'cfvo', [])
        if cfvo_list:
            cs_data['cfvo'] = [
                {k: v for k, v in [
                    ('type', getattr(cfvo, 'type', None)),
                    ('val', getattr(cfvo, 'val', None)),
                ] if v is not None}
                for cfvo in cfvo_list
            ]
        colors = getattr(color_scale, 'color', [])
        if colors:
            color_vals = []
            for c in colors:
                extracted = _extract_color_raw(c)
                if extracted is not None:
                    color_vals.append(extracted)
            if color_vals:
                cs_data['color'] = color_vals
        result['colorScale'] = cs_data

    # dataBar
    data_bar = getattr(rule, 'dataBar', None)
    if data_bar is not None:
        db_data: dict = {}
        cfvo_list = getattr(data_bar, 'cfvo', [])
        if cfvo_list:
            db_data['cfvo'] = [
                {k: v for k, v in [
                    ('type', getattr(cfvo, 'type', None)),
                    ('val', getattr(cfvo, 'val', None)),
                ] if v is not None}
                for cfvo in cfvo_list
            ]
        color = getattr(data_bar, 'color', None)
        if color is not None:
            extracted = _extract_color_raw(color)
            if extracted is not None:
                db_data['color'] = extracted
        result['dataBar'] = db_data

    # iconSet
    icon_set = getattr(rule, 'iconSet', None)
    if icon_set is not None:
        is_data: dict = {}
        cfvo_list = getattr(icon_set, 'cfvo', [])
        if cfvo_list:
            is_data['cfvo'] = [
                {k: v for k, v in [
                    ('type', getattr(cfvo, 'type', None)),
                    ('val', getattr(cfvo, 'val', None)),
                ] if v is not None}
                for cfvo in cfvo_list
            ]
        icon_set_str = getattr(icon_set, 'iconSet', None)
        if icon_set_str:
            is_data['iconSet'] = icon_set_str
        result['iconSet'] = is_data

    return result


def export_conditional_formatting_json(wb_data: openpyxl.Workbook, output_dir: Path) -> Path:
    """
    全シートの条件付き書式を conditional_formatting.json に出力。
    ws.conditional_formatting._cf_rules.items() を走査。

    出力形式:
    {
      "シート名": [
        {
          "sqref": "G10:H10",
          "rules": [
            {
              "type": "cellIs",
              "operator": "greaterThan",
              "formula": ["0"],
              "priority": 1,
              "dxf": {...}
            }
          ]
        }
      ]
    }
    """
    cf_data: dict[str, list] = {}

    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]
        sheet_cf: list = []

        # _cf_rules: { ConditionalFormatting(key) -> list[Rule] }
        # key.sqref が MultiCellRange オブジェクト（str() で "A1:B2 C3:D4" 形式）
        for cf_key, rules in ws.conditional_formatting._cf_rules.items():
            sqref_str = str(cf_key.sqref)
            rule_list = []
            for rule in rules:
                serialized = _serialize_cf_rule(rule)
                if serialized:
                    rule_list.append(serialized)
            if rule_list:
                sheet_cf.append({
                    'sqref': sqref_str,
                    'rules': rule_list,
                })

        if sheet_cf:
            cf_data[sheet_name] = sheet_cf

    cf_path = output_dir / 'conditional_formatting.json'
    with open(cf_path, 'w', encoding='utf-8') as f:
        json.dump(cf_data, f, ensure_ascii=False, indent=2)

    return cf_path


# ---- data_validation.json 出力 ----------------------------------------------

def export_data_validation_json(wb_data: openpyxl.Workbook, output_dir: Path) -> Path:
    """
    全シートのデータバリデーション（プルダウン・入力規則）を
    data_validation.json に出力。

    出力形式:
    {
      "シート名": [
        {
          "type": "list",
          "formula1": "'エネミー出現'!$M$2:$M61",
          "formula2": null,
          "operator": null,
          "allow_blank": true,
          "showDropDown": false,
          "showInputMessage": false,
          "showErrorMessage": false,
          "errorTitle": null,
          "error": null,
          "promptTitle": null,
          "prompt": null,
          "sqref": "M10:M86"
        }
      ]
    }
    """
    dv_data: dict[str, list] = {}

    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]
        sheet_dvs: list = []

        for dv in ws.data_validations.dataValidation:
            entry = {
                'type': getattr(dv, 'type', None),
                'formula1': getattr(dv, 'formula1', None),
                'formula2': getattr(dv, 'formula2', None),
                'operator': getattr(dv, 'operator', None),
                'allow_blank': getattr(dv, 'allow_blank', True),
                'showDropDown': getattr(dv, 'showDropDown', False),
                'showInputMessage': getattr(dv, 'showInputMessage', False),
                'showErrorMessage': getattr(dv, 'showErrorMessage', False),
                'errorTitle': getattr(dv, 'errorTitle', None),
                'error': getattr(dv, 'error', None),
                'promptTitle': getattr(dv, 'promptTitle', None),
                'prompt': getattr(dv, 'prompt', None),
                'sqref': str(dv.sqref),
            }
            sheet_dvs.append(entry)

        if sheet_dvs:
            dv_data[sheet_name] = sheet_dvs

    dv_path = output_dir / 'data_validation.json'
    with open(dv_path, 'w', encoding='utf-8') as f:
        json.dump(dv_data, f, ensure_ascii=False, indent=2)

    return dv_path


# ---- CSV 出力 ---------------------------------------------------------------

def export_csv(wb_data: openpyxl.Workbook, wb_formula: openpyxl.Workbook, output_dir: Path) -> list[dict]:
    """
    全シートの実データを CSV に出力する。

    IMPORTRANGE で転記されたセル (data_only=True で None になる) は
    式の IFERROR フォールバックから値を補完する。
    """
    csv_dir = output_dir / 'csv'
    csv_dir.mkdir(exist_ok=True)

    sheet_info = []

    for sheet_name in wb_data.sheetnames:
        ws_data = wb_data[sheet_name]
        ws_form = wb_formula[sheet_name]
        safe = sanitize_filename(sheet_name)
        csv_path = csv_dir / f'{safe}.csv'

        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row_data, row_form in zip(ws_data.iter_rows(), ws_form.iter_rows()):
                out_row = []
                for cell_d, cell_f in zip(row_data, row_form):
                    value = cell_d.value

                    # data_only=True で None & 式セルの場合 → フォールバックを補完
                    if value is None and cell_f.data_type == 'f' and isinstance(cell_f.value, str):
                        info = classify_formula(cell_f.value)
                        if info['formula_type'] == 'computed_value':
                            value = info.get('value')  # IMPORTRANGE 転記値
                        elif info['formula_type'] == 'importrange':
                            value = None  # ヘッダー行はそのまま空
                        # 通常式は None のまま（キャッシュなし）

                    out_row.append(cell_value_to_str(value))

                # 末尾の空列を除去してから書き込み
                while out_row and out_row[-1] == '':
                    out_row.pop()
                writer.writerow(out_row)

        sheet_info.append({
            'sheet_name': sheet_name,
            'file': f'csv/{safe}.csv',
            'max_row': ws_data.max_row,
            'max_col': ws_data.max_column,
        })

    return sheet_info


# ---- cells.json 出力 --------------------------------------------------------

def export_cells_json(wb_formula: openpyxl.Workbook, output_dir: Path) -> Path:
    """
    式を持つセルの情報を cells.json に出力する。

    出力形式:
    {
      "シート名": [
        {
          "coord": "A1",
          "row": 1,
          "col": 1,
          "col_letter": "A",
          "formula_type": "importrange" | "computed_value" | "normal",
          "formula": "=...",
          // formula_type="importrange" の場合
          "importrange_url": "https://...",
          "importrange_range": "シート!A:N",
          // formula_type="computed_value" の場合
          "value": "実際の値",
        },
        ...
      ]
    }
    """
    cells_data: dict[str, list] = {}

    for sheet_name in wb_formula.sheetnames:
        ws = wb_formula[sheet_name]
        sheet_cells = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != 'f':
                    continue  # 式セルのみ対象

                # ArrayFormula オブジェクトと通常文字列を両方処理
                if isinstance(cell.value, ArrayFormula):
                    formula_text = cell.value.text
                    is_array = True
                elif isinstance(cell.value, str):
                    formula_text = cell.value
                    is_array = False
                else:
                    continue  # 未知の型はスキップ

                if not formula_text:
                    continue  # formula_text が None or 空文字はスキップ

                info = classify_formula(formula_text)
                entry = {
                    'coord': cell.coordinate,
                    'row': cell.row,
                    'col': cell.column,
                    'col_letter': get_column_letter(cell.column),
                    **info,
                }
                if is_array:
                    entry['is_array_formula'] = True
                sheet_cells.append(entry)

        if sheet_cells:
            cells_data[sheet_name] = sheet_cells

    cells_path = output_dir / 'cells.json'
    with open(cells_path, 'w', encoding='utf-8') as f:
        json.dump(cells_data, f, ensure_ascii=False, indent=2)

    return cells_path


# ---- metadata.json 出力 -----------------------------------------------------

def export_metadata(xlsx_path: Path, sheet_info: list[dict], cells_path: Path, output_dir: Path) -> Path:
    """ブック全体のメタ情報を metadata.json に出力する"""
    # シート別の式タイプ集計（cells.json から）
    with open(cells_path, encoding='utf-8') as f:
        cells_data = json.load(f)

    formula_summary = {}
    for sname, cells in cells_data.items():
        types: dict[str, int] = {}
        for c in cells:
            t = c.get('formula_type', 'normal')
            types[t] = types.get(t, 0) + 1
        formula_summary[sname] = types

    metadata = {
        'source_file': xlsx_path.name,
        'converted_at': datetime.now().isoformat(),
        'total_sheets': len(sheet_info),
        'sheets': [
            {
                **s,
                'formula_counts': formula_summary.get(s['sheet_name'], {}),
            }
            for s in sheet_info
        ],
    }

    meta_path = output_dir / 'metadata.json'
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    return meta_path


# ---- メイン -----------------------------------------------------------------

def convert(xlsx_path: Path, output_dir: Path) -> None:
    print(f'変換開始: {xlsx_path.name}')

    print('  [1/8] 実データ読み込み (data_only=True)...')
    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)

    print('  [2/8] 式情報読み込み (data_only=False)...')
    wb_formula = openpyxl.load_workbook(xlsx_path, data_only=False)

    output_dir.mkdir(parents=True, exist_ok=True)

    print('  [3/8] CSV 出力中...')
    sheet_info = export_csv(wb_data, wb_formula, output_dir)
    print(f'         {len(sheet_info)} シートを出力しました')

    print('  [4/8] cells.json / metadata.json 出力中...')
    cells_path = export_cells_json(wb_formula, output_dir)
    meta_path = export_metadata(xlsx_path, sheet_info, cells_path, output_dir)

    print('  [5/8] styles.json 出力中...')
    styles_path = export_styles_json(wb_data, output_dir)
    total_style_cells = sum(
        len(cells)
        for cells in json.load(open(styles_path, encoding='utf-8')).values()
    )
    print(f'         {total_style_cells:,} セルのスタイル情報を出力しました')

    print('  [6/8] layout.json 出力中...')
    layout_path = export_layout_json(wb_data, output_dir)
    layout_data = json.load(open(layout_path, encoding='utf-8'))
    total_merged = sum(
        len(s.get('merged_cells', [])) for s in layout_data.values()
    )
    sheets_with_freeze = sum(
        1 for s in layout_data.values() if s.get('freeze_panes')
    )
    print(f'         結合セル: {total_merged:,} 個、フリーズペイン: {sheets_with_freeze} シート')

    print('  [7/8] conditional_formatting.json 出力中...')
    cf_path = export_conditional_formatting_json(wb_data, output_dir)
    cf_data = json.load(open(cf_path, encoding='utf-8'))
    total_cf_entries = sum(len(entries) for entries in cf_data.values())
    print(f'         {total_cf_entries} エントリの条件付き書式を出力しました')

    print('  [8/8] data_validation.json 出力中...')
    dv_path = export_data_validation_json(wb_formula, output_dir)
    dv_data = json.load(open(dv_path, encoding='utf-8'))
    total_dv_entries = sum(len(entries) for entries in dv_data.values())
    print(f'         {total_dv_entries} 個のデータバリデーションを出力しました')

    print(f'\n完了: {output_dir}')
    print(f'  csv/                          : {len(sheet_info)} ファイル')
    print(f'  cells.json                    : {cells_path.stat().st_size:,} bytes')
    print(f'  metadata.json                 : {meta_path.stat().st_size:,} bytes')
    print(f'  styles.json                   : {styles_path.stat().st_size:,} bytes')
    print(f'  layout.json                   : {layout_path.stat().st_size:,} bytes')
    print(f'  conditional_formatting.json   : {cf_path.stat().st_size:,} bytes')
    print(f'  data_validation.json          : {dv_path.stat().st_size:,} bytes')


def main() -> None:
    parser = argparse.ArgumentParser(
        description='XLSX → CSV（実データ）+ JSON（式情報）変換ツール'
    )
    parser.add_argument('xlsx_file', type=Path, help='変換する XLSX ファイルのパス')
    parser.add_argument(
        '--output-dir', '-o', type=Path,
        help='出力ディレクトリ（省略時: outputs/<ファイル名>/ に自動生成）',
    )
    args = parser.parse_args()

    xlsx_path = args.xlsx_file.resolve()
    if not xlsx_path.exists():
        print(f'エラー: ファイルが見つかりません: {xlsx_path}', file=sys.stderr)
        sys.exit(1)
    if xlsx_path.suffix.lower() != '.xlsx':
        print(f'エラー: .xlsx ファイルを指定してください', file=sys.stderr)
        sys.exit(1)

    if args.output_dir:
        output_dir = args.output_dir.resolve()
    else:
        # デフォルト: スクリプトの1つ上の outputs/<ファイル名>/ フォルダ
        output_dir = xlsx_path.parent.parent / 'outputs' / xlsx_path.stem

    convert(xlsx_path, output_dir)


if __name__ == '__main__':
    main()
