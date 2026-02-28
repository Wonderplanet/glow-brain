"""
MstPage/MstKomaLine投入用の汎用セル式を全対象シートに適用するスクリプト。

汎用化の仕組み:
- MATCH("敵ゲートID",E:E,0)+1 でpage_idのデータ行を動的取得
- MATCH("リリースキー",N:N,0)+1 でrelease_keyのデータ行を動的取得
- MATCH("■コマ設計",B:B,0)+3+(i-1) でコマ行データ行を動的取得
- LET関数で式内変数化（Google Sheets 2022年以降対応）

対象:
- 【チャレンジ】死罪人と首切り役人設計.xlsx: 1話〜4話
- 【降臨バトル】まるで 悪夢を見ているようだ_地獄楽.xlsx: 1話
- 【高難度】手負いの獣は恐ろしいぞ.xlsx: 1話〜3話
- 検証用コピー_未完成_限界チャレンジ(VD)_アウトゲーム関連.xlsx: 通常ブロック, ボスブロック
- 検証用コピー_【ストーリー】必ず生きて帰る.xlsx: 1話〜6話（既存式を動的式に更新）
"""
import openpyxl
from openpyxl.utils import column_index_from_string
import shutil
import os

TASK_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(TASK_DIR, 'raw')
WORK_DIR = os.path.join(TASK_DIR, 'work')

# MstKomaLineのカラムヘッダー（EB〜FRの43列）
KOMA_LINE_HEADERS = [
    'ENABLE', 'id', 'mst_page_id', 'row', 'height', 'koma_line_layout_asset_key',
    'koma1_asset_key', 'koma1_width', 'koma1_back_ground_offset',
    'koma1_effect_type', 'koma1_effect_parameter1', 'koma1_effect_parameter2',
    'koma1_effect_target_side', 'koma1_effect_target_colors', 'koma1_effect_target_roles',
    'koma2_asset_key', 'koma2_width', 'koma2_back_ground_offset',
    'koma2_effect_type', 'koma2_effect_parameter1', 'koma2_effect_parameter2',
    'koma2_effect_target_side', 'koma2_effect_target_colors', 'koma2_effect_target_roles',
    'koma3_asset_key', 'koma3_width', 'koma3_back_ground_offset',
    'koma3_effect_type', 'koma3_effect_parameter1', 'koma3_effect_parameter2',
    'koma3_effect_target_side', 'koma3_effect_target_colors', 'koma3_effect_target_roles',
    'koma4_asset_key', 'koma4_width', 'koma4_back_ground_offset',
    'koma4_effect_type', 'koma4_effect_parameter1', 'koma4_effect_parameter2',
    'koma4_effect_target_side', 'koma4_effect_target_colors', 'koma4_effect_target_roles',
    'release_key',
]

EB_COL = column_index_from_string('EB')


def gen_page_formulas():
    """MstPage用の式（29行目）を生成"""
    return {
        'EB': '=LET(er,MATCH("敵ゲートID",E:E,0)+1,IF(INDIRECT("E"&er)="","","e"))',
        'EC': '=LET(er,MATCH("敵ゲートID",E:E,0)+1,IF(INDIRECT("E"&er)="","",INDIRECT("E"&er)))',
        'ED': ('=LET(er,MATCH("敵ゲートID",E:E,0)+1,'
               'rr,MATCH("リリースキー",N:N,0)+1,'
               'IF(INDIRECT("E"&er)="","",TEXT(INDIRECT("N"&rr),"0")))'),
    }


def gen_komaline_formulas(row_num):
    """MstKomaLine用の式を行番号（1-5）ごとに生成"""
    offset = row_num - 1  # 1行目=+3, 2行目=+4, ...
    kr_expr = f'MATCH("■コマ設計",B:B,0)+{3 + offset}'
    er_expr = 'MATCH("敵ゲートID",E:E,0)+1'
    rr_expr = 'MATCH("リリースキー",N:N,0)+1'

    # LETラッパー
    def let_kr(inner):
        return f'=LET(kr,{kr_expr},{inner})'

    def let_er_kr(inner):
        return f'=LET(er,{er_expr},kr,{kr_expr},{inner})'

    def let_all(inner):
        return f'=LET(er,{er_expr},rr,{rr_expr},kr,{kr_expr},{inner})'

    def if_d(then, else_='""'):
        return f'IF(INDIRECT("D"&kr)="",{else_},{then})'

    def koma_exists(w):
        return (f'AND({w}<>"",{w}<>"none",{w}<>"error",'
                f'NOT(ISBLANK({w})))')

    return {
        # --- 基本フィールド ---
        'EB': let_kr(if_d('"e"')),
        'EC': let_er_kr(if_d(f'INDIRECT("E"&er)&"_{row_num}"')),
        'ED': let_er_kr(if_d('INDIRECT("E"&er)')),
        'EE': let_kr(if_d(str(row_num))),
        'EF': let_kr(if_d('INDIRECT("H"&kr)')),
        'EG': let_kr(if_d('INT(INDIRECT("D"&kr))')),
        # --- koma1 ---
        'EH': let_kr(if_d('IFERROR(INDIRECT("R"&kr),"")')),
        'EI': let_kr(if_d('INDIRECT("J"&kr)')),
        'EJ': let_kr(if_d('IF(INDIRECT("J"&kr)=1,0,-1)')),
        'EK': let_kr(if_d(
            'IF(OR(INDIRECT("U"&kr)="",ISBLANK(INDIRECT("U"&kr))),"None",INDIRECT("U"&kr))')),
        'EL': let_kr(if_d(
            'IF(OR(INDIRECT("U"&kr)="",ISBLANK(INDIRECT("U"&kr))),0,INDIRECT("Z"&kr))')),
        'EM': let_kr(if_d(
            'IF(OR(INDIRECT("U"&kr)="",ISBLANK(INDIRECT("U"&kr))),0,INDIRECT("AB"&kr))')),
        'EN': let_kr(if_d('"All"')),
        'EO': let_kr(if_d('"All"')),
        'EP': let_kr(if_d('"All"')),
        # --- koma2 ---
        'EQ': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},'
            f'IFERROR(INDIRECT("R"&kr),""),"")')),
        'ER': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},'
            f'INDIRECT("L"&kr),"")')),
        'ES': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},'
            f'IF(INDIRECT("L"&kr)=1,0,-1),"__NULL__")')),
        'ET': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},'
            f'IF(OR(INDIRECT("AD"&kr)="",ISBLANK(INDIRECT("AD"&kr))),'
            f'"None",INDIRECT("AD"&kr)),"None")')),
        'EU': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},'
            f'IF(OR(INDIRECT("AD"&kr)="",ISBLANK(INDIRECT("AD"&kr))),'
            f'0,INDIRECT("AI"&kr)),"")')),
        'EV': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},'
            f'IF(OR(INDIRECT("AD"&kr)="",ISBLANK(INDIRECT("AD"&kr))),'
            f'0,INDIRECT("AK"&kr)),"")')),
        'EW': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},"All","__NULL__")')),
        'EX': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},"All","")')),
        'EY': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"L\"&kr)")},"All","")')),
        # --- koma3 ---
        'EZ': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},'
            f'IFERROR(INDIRECT("R"&kr),""),"")')),
        'FA': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},'
            f'INDIRECT("N"&kr),"")')),
        'FB': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},'
            f'IF(INDIRECT("N"&kr)=1,0,-1),"__NULL__")')),
        'FC': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},'
            f'IF(OR(INDIRECT("AM"&kr)="",ISBLANK(INDIRECT("AM"&kr))),'
            f'"None",INDIRECT("AM"&kr)),"None")')),
        'FD': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},'
            f'IF(OR(INDIRECT("AM"&kr)="",ISBLANK(INDIRECT("AM"&kr))),'
            f'0,INDIRECT("AR"&kr)),"")')),
        'FE': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},'
            f'IF(OR(INDIRECT("AM"&kr)="",ISBLANK(INDIRECT("AM"&kr))),'
            f'0,INDIRECT("AT"&kr)),"")')),
        'FF': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},"All","__NULL__")')),
        'FG': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},"All","")')),
        'FH': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"N\"&kr)")},"All","")')),
        # --- koma4 ---
        'FI': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},'
            f'IFERROR(INDIRECT("R"&kr),""),"")')),
        'FJ': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},'
            f'INDIRECT("P"&kr),"")')),
        'FK': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},'
            f'IF(INDIRECT("P"&kr)=1,0,-1),"__NULL__")')),
        'FL': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},'
            f'IF(OR(INDIRECT("AV"&kr)="",ISBLANK(INDIRECT("AV"&kr))),'
            f'"None",INDIRECT("AV"&kr)),"None")')),
        'FM': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},'
            f'IF(OR(INDIRECT("AV"&kr)="",ISBLANK(INDIRECT("AV"&kr))),'
            f'0,INDIRECT("BA"&kr)),"")')),
        'FN': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},'
            f'IF(OR(INDIRECT("AV"&kr)="",ISBLANK(INDIRECT("AV"&kr))),'
            f'0,INDIRECT("BC"&kr)),"")')),
        'FO': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},"All","__NULL__")')),
        'FP': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},"All","")')),
        'FQ': let_kr(if_d(
            f'IF({koma_exists("INDIRECT(\"P\"&kr)")},"All","")')),
        # --- release_key ---
        'FR': let_all(
            'IF(INDIRECT("D"&kr)="","",TEXT(INDIRECT("N"&rr),"0"))'),
    }


def apply_formulas_to_sheet(ws):
    """指定シートに★MstPage/MstKomaLine投入用ヘッダー・式を書き込む"""
    # EA28: ★MstPage投入用ヘッダー
    ws['EA28'] = '★MstPage投入用▶'
    ws['EB28'] = 'ENABLE'
    ws['EC28'] = 'id'
    ws['ED28'] = 'release_key'

    # EB29〜ED29: MstPageデータ式
    for col_letter, formula in gen_page_formulas().items():
        ws[f'{col_letter}29'] = formula

    # EA30: ★MstKomaLine投入用ヘッダー
    ws['EA30'] = '★MstKomaLine投入用▶'

    # EB30〜FR30: カラムヘッダー
    for i, header in enumerate(KOMA_LINE_HEADERS):
        ws.cell(row=30, column=EB_COL + i, value=header)

    # EB31〜FR35: 1〜5行目データ式
    for row_num in range(1, 6):
        excel_row = 30 + row_num
        for col_letter, formula in gen_komaline_formulas(row_num).items():
            ws.cell(row=excel_row, column=column_index_from_string(col_letter),
                    value=formula)

    print(f'  ✓ {ws.title}')


def ensure_work_copy(raw_filename, work_filename=None):
    if work_filename is None:
        work_filename = raw_filename
    raw_path = os.path.join(RAW_DIR, raw_filename)
    work_path = os.path.join(WORK_DIR, work_filename)
    if not os.path.exists(work_path):
        shutil.copy2(raw_path, work_path)
        print(f'  rawからworkにコピー: {work_filename}')
    return work_path


def process_file(work_path, target_sheets):
    print(f'\n[{os.path.basename(work_path)}]')
    wb = openpyxl.load_workbook(work_path)
    for sheet_name in target_sheets:
        if sheet_name in wb.sheetnames:
            apply_formulas_to_sheet(wb[sheet_name])
        else:
            print(f'  ⚠ シートが見つかりません: {sheet_name}')
    wb.save(work_path)
    print(f'  → 保存完了')


def main():
    os.makedirs(WORK_DIR, exist_ok=True)

    targets = [
        ('【チャレンジ】死罪人と首切り役人設計.xlsx',
         '【チャレンジ】死罪人と首切り役人設計.xlsx',
         ['1話', '2話', '3話', '4話']),
        ('【降臨バトル】まるで 悪夢を見ているようだ_地獄楽.xlsx',
         '【降臨バトル】まるで 悪夢を見ているようだ_地獄楽.xlsx',
         ['1話']),
        ('【高難度】手負いの獣は恐ろしいぞ.xlsx',
         '【高難度】手負いの獣は恐ろしいぞ.xlsx',
         ['1話', '2話', '3話']),
        ('検証用コピー_未完成_限界チャレンジ(VD)_アウトゲーム関連.xlsx',
         '検証用コピー_未完成_限界チャレンジ(VD)_アウトゲーム関連.xlsx',
         ['通常ブロック', 'ボスブロック']),
        # ストーリーはwork/に既存 → 式を上書き更新
        ('検証用コピー_【ストーリー】必ず生きて帰る.xlsx',
         '検証用コピー_【ストーリー】必ず生きて帰る.xlsx',
         ['1話', '2話', '3話', '4話', '5話', '6話']),
    ]

    for raw_name, work_name, sheets in targets:
        work_path = ensure_work_copy(raw_name, work_name)
        process_file(work_path, sheets)

    print('\n=== 全ファイル処理完了 ===')


if __name__ == '__main__':
    main()
