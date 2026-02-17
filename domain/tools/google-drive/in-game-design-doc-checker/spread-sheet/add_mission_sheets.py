"""
20260116_地獄楽 いいジャン祭_仕様書.xlsx に
イベントミッション用GAS設計書シートを7枚追加するスクリプト。

追加シート:
  1. チェック対象管理
  2. ミッション一覧
  3. 依存関係一覧
  4. 報酬一覧
  5. マッピング設定_ミッション
  6. マッピング設定_依存関係
  7. マッピング設定_報酬
"""

import csv
import io
import openpyxl
from pathlib import Path

# ファイルパス
XLSX_PATH = Path(__file__).parent / "20260116_地獄楽 いいジャン祭_仕様書.xlsx"
CSV_DIR = Path(__file__).parents[4] / "raw-data/masterdata/released/202601010/tables"

# CSVデータ読み込み
def read_csv(name: str) -> list[list]:
    path = CSV_DIR / f"{name}.csv"
    rows = []
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    # 末尾の空行を除去
    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    return rows


# -------------------------------------------------
# 各CSVを読み込む
# -------------------------------------------------
mission_rows = read_csv("MstMissionEvent")           # ヘッダー含む
i18n_rows    = read_csv("MstMissionEventI18n")       # ヘッダー含む
dep_rows     = read_csv("MstMissionEventDependency") # ヘッダー含む
reward_rows  = read_csv("MstMissionReward")          # ヘッダー含む

# 報酬は jig_00001_event_reward_* のみ
reward_data_rows = [r for r in reward_rows[1:] if r[3].startswith("jig_00001_event_reward_")]

# -------------------------------------------------
# xlsxを開く
# -------------------------------------------------
wb = openpyxl.load_workbook(XLSX_PATH)

EXISTING = set(wb.sheetnames)
print(f"既存シート数: {len(EXISTING)}")

# シートを末尾に追加するヘルパー
def add_sheet(name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in EXISTING:
        print(f"  既存シートを削除して再作成: {name}")
        del wb[name]
    ws = wb.create_sheet(name)
    print(f"  追加: {name}")
    return ws


# =================================================
# 1. チェック対象管理
# =================================================
ws = add_sheet("チェック対象管理")

header = ["チェック", "シート名", "マッピング", "マスターURL", "最終チェック日時", "結果"]
ws.append(header)

check_rows = [
    [True,  "ミッション一覧",  "マッピング設定_ミッション",   "", "", ""],
    [True,  "依存関係一覧",    "マッピング設定_依存関係",     "", "", ""],
    [True,  "報酬一覧",        "マッピング設定_報酬",         "", "", ""],
]
for row in check_rows:
    ws.append(row)


# =================================================
# 2. ミッション一覧
# -------------------------------------------------
# Row1: タイトル, Row2: ヘッダー, Row3〜: データ
# 列A〜N: MstMissionEvent、列O〜Q: MstMissionEventI18n
# =================================================
ws = add_sheet("ミッション一覧")

# Row1: タイトル
ws.append(["ミッション一覧（MstMissionEvent + MstMissionEventI18n）"])

# Row2: ヘッダー
ws.append([
    "ENABLE", "id", "release_key", "mst_event_id",
    "criterion_type", "criterion_value", "criterion_count",
    "unlock_criterion_type", "unlock_criterion_value", "unlock_criterion_count",
    "group_key", "mst_mission_reward_group_id", "sort_order", "destination_scene",
    "i18n_id", "language", "description",
])

# i18n を id順に辞書化
# i18n CSVヘッダー: ENABLE, release_key, id, mst_mission_event_id, language, description
i18n_dict = {}
for r in i18n_rows[1:]:
    # r[3] = mst_mission_event_id
    i18n_dict[r[3]] = r  # キー = mst_mission_event_id

# MstMissionEvent データ行 (Row3〜)
# CSVヘッダー: ENABLE,id,release_key,mst_event_id,criterion_type,criterion_value,criterion_count,
#              unlock_criterion_type,unlock_criterion_value,unlock_criterion_count,
#              group_key,mst_mission_reward_group_id,sort_order,destination_scene
for m_row in mission_rows[1:]:
    mission_id = m_row[1]  # id列
    i_row = i18n_dict.get(mission_id, ["", "", "", "", "", ""])
    # i18n列: i18n_id=i_row[2], language=i_row[4], description=i_row[5]
    full_row = m_row[:14] + [i_row[2], i_row[4], i_row[5]]
    ws.append(full_row)


# =================================================
# 3. 依存関係一覧
# =================================================
ws = add_sheet("依存関係一覧")

ws.append(["依存関係一覧（MstMissionEventDependency）"])
ws.append(["ENABLE", "id", "release_key", "group_id", "mst_mission_event_id", "unlock_order", "備考"])

for row in dep_rows[1:]:
    ws.append(row)


# =================================================
# 4. 報酬一覧
# =================================================
ws = add_sheet("報酬一覧")

ws.append(["報酬一覧（MstMissionReward - イベントミッション分）"])
ws.append(["ENABLE", "id", "release_key", "group_id", "resource_type", "resource_id", "resource_amount", "sort_order", "備考"])

for row in reward_data_rows:
    ws.append(row)


# =================================================
# 5. マッピング設定_ミッション
# =================================================
ws = add_sheet("マッピング設定_ミッション")

# ヘッダー行
mapping_header = ["項目名", "方向", "設計書範囲", "マスター名", "マスター列名", "チェック無視", "ユニークキー", "展開タイプ", "空白無視"]
ws.append(mapping_header)

# MstMissionEvent 用マッピング (14行)
mission_event_mappings = [
    ["", "", "A3:A100", "MstMissionEvent", "ENABLE",                       False, False, "通常", False],
    ["", "", "B3:B100", "MstMissionEvent", "id",                           False, True,  "通常", False],
    ["", "", "C3:C100", "MstMissionEvent", "release_key",                  False, False, "通常", False],
    ["", "", "D3:D100", "MstMissionEvent", "mst_event_id",                 False, False, "通常", False],
    ["", "", "E3:E100", "MstMissionEvent", "criterion_type",               False, False, "通常", False],
    ["", "", "F3:F100", "MstMissionEvent", "criterion_value",              False, False, "通常", False],
    ["", "", "G3:G100", "MstMissionEvent", "criterion_count",              False, False, "通常", False],
    ["", "", "H3:H100", "MstMissionEvent", "unlock_criterion_type",       False, False, "通常", False],
    ["", "", "I3:I100", "MstMissionEvent", "unlock_criterion_value",      False, False, "通常", False],
    ["", "", "J3:J100", "MstMissionEvent", "unlock_criterion_count",      False, False, "通常", False],
    ["", "", "K3:K100", "MstMissionEvent", "group_key",                   False, False, "通常", False],
    ["", "", "L3:L100", "MstMissionEvent", "mst_mission_reward_group_id", False, False, "通常", False],
    ["", "", "M3:M100", "MstMissionEvent", "sort_order",                  False, False, "通常", False],
    ["", "", "N3:N100", "MstMissionEvent", "destination_scene",           False, False, "通常", False],
]

# MstMissionEventI18n 用マッピング (6行)
mission_i18n_mappings = [
    ["", "", "A3:A100",  "MstMissionEventI18n", "ENABLE",               False, False, "通常", False],
    ["", "", "C3:C100",  "MstMissionEventI18n", "release_key",          False, False, "通常", False],
    ["", "", "O3:O100",  "MstMissionEventI18n", "id",                   False, True,  "通常", False],
    ["", "", "B3:B100",  "MstMissionEventI18n", "mst_mission_event_id", False, False, "通常", False],
    ["", "", "P3:P100",  "MstMissionEventI18n", "language",             False, False, "通常", False],
    ["", "", "Q3:Q100",  "MstMissionEventI18n", "description",          False, False, "通常", False],
]

for row in mission_event_mappings + mission_i18n_mappings:
    ws.append(row)


# =================================================
# 6. マッピング設定_依存関係
# =================================================
ws = add_sheet("マッピング設定_依存関係")

ws.append(["項目名", "方向", "設計書範囲", "マスター名", "マスター列名", "チェック無視", "ユニークキー", "展開タイプ", "空白無視"])

dep_mappings = [
    ["", "", "A3:A100", "MstMissionEventDependency", "ENABLE",               False, False, "通常", False],
    ["", "", "B3:B100", "MstMissionEventDependency", "id",                   False, True,  "通常", False],
    ["", "", "C3:C100", "MstMissionEventDependency", "release_key",          False, False, "通常", False],
    ["", "", "D3:D100", "MstMissionEventDependency", "group_id",             False, False, "通常", False],
    ["", "", "E3:E100", "MstMissionEventDependency", "mst_mission_event_id", False, False, "通常", False],
    ["", "", "F3:F100", "MstMissionEventDependency", "unlock_order",         False, False, "通常", False],
    ["", "", "G3:G100", "MstMissionEventDependency", "備考",                 False, False, "通常", False],
]

for row in dep_mappings:
    ws.append(row)


# =================================================
# 7. マッピング設定_報酬
# =================================================
ws = add_sheet("マッピング設定_報酬")

ws.append(["項目名", "方向", "設計書範囲", "マスター名", "マスター列名", "チェック無視", "ユニークキー", "展開タイプ", "空白無視"])

reward_mappings = [
    ["", "", "A3:A100", "MstMissionReward", "ENABLE",          False, False, "通常", False],
    ["", "", "B3:B100", "MstMissionReward", "id",              False, True,  "通常", False],
    ["", "", "C3:C100", "MstMissionReward", "release_key",     False, False, "通常", False],
    ["", "", "D3:D100", "MstMissionReward", "group_id",        False, False, "通常", False],
    ["", "", "E3:E100", "MstMissionReward", "resource_type",   False, False, "通常", False],
    ["", "", "F3:F100", "MstMissionReward", "resource_id",     False, False, "通常", False],
    ["", "", "G3:G100", "MstMissionReward", "resource_amount", False, False, "通常", False],
    ["", "", "H3:H100", "MstMissionReward", "sort_order",      False, False, "通常", False],
    ["", "", "I3:I100", "MstMissionReward", "備考",            False, False, "通常", False],
]

for row in reward_mappings:
    ws.append(row)


# =================================================
# 保存
# =================================================
wb.save(XLSX_PATH)
print(f"\n保存完了: {XLSX_PATH}")

# 検証
wb2 = openpyxl.load_workbook(XLSX_PATH)
print(f"\n最終シート数: {len(wb2.sheetnames)}")

added = ["チェック対象管理", "ミッション一覧", "依存関係一覧", "報酬一覧",
         "マッピング設定_ミッション", "マッピング設定_依存関係", "マッピング設定_報酬"]
for name in added:
    ws = wb2[name]
    # データ行数 (Row1=タイトル or ヘッダー, Row2=ヘッダー or データ開始)
    max_row = ws.max_row
    print(f"  {name}: max_row={max_row}")
